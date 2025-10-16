"""
Daily Tracker Manager
Handles Excel file operations for daily trackers
"""
import os
import shutil
from datetime import datetime, date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal

# Tracker type mappings
TRACKER_TYPES = {
    '50K': '50K Reinvest Tracker',
    '1L': '1L Enhanced Tracker',
    'No Reinvest': 'No Reinvest Tracker'
}

# Sheet column configurations for different tracker types
SHEET_CONFIGS = {
    '50K': {
        'sheet_name': '50K Reinvest Tracker',
        'tracker_name_cell': 'B3',
        'investment_cell': 'B6',
        'scheme_period_cell': 'B7',
        'start_date_cell': 'B10',
        'per_day_payment_cell': 'B8',
        'data_start_row': 13,
        'columns': {
            'day': 'A',
            'date': 'B',
            'daily_payments': 'C',
            'payment_mode': 'E',
            'cumulative': 'D',
            'balance': 'F',
            'withdrawn': 'G',
            'notes': 'H'
        }
    },
    '1L': {
        'sheet_name': '1L Enhanced Tracker',
        'tracker_name_cell': 'B3',
        'investment_cell': 'B6',
        'scheme_period_cell': 'B7',
        'start_date_cell': 'B11',
        'per_day_payment_cell': 'B8',
        'data_start_row': 14,
        'columns': {
            'day': 'A',
            'date': 'B',
            'daily_payments': 'C',
            'payment_mode': 'E',
            'cumulative': 'D',
            'balance': 'F',
            'withdrawn': 'G',
            'reinvest': 'H',
            'reinvest_from': 'I',
            'pocket_money': 'J',
            'total_invested': 'K',
            'notes': 'L'
        }
    },
    'No Reinvest': {
        'sheet_name': 'No Reinvest Tracker',
        'tracker_name_cell': 'B3',
        'investment_cell': 'B6',
        'scheme_period_cell': 'B7',
        'start_date_cell': 'B10',
        'per_day_payment_cell': 'B8',
        'data_start_row': 13,
        'columns': {
            'day': 'A',
            'date': 'B',
            'daily_payments': 'C',
            'transaction_details': 'D',
            'payment_mode': 'E',
            'cumulative': 'F',
            'balance': 'G',
            'withdrawn': 'H',
            'notes': 'I'
        }
    }
}


def get_tracker_directory(instance):
    """Get the daily tracker directory for a specific instance"""
    base_dir = Path("instances") / instance / "daily-trackers"
    base_dir.mkdir(parents=True, exist_ok=True)
    return str(base_dir)


def create_tracker_file(instance, username, tracker_name, tracker_type, investment, scheme_period, start_date, per_day_payment=None):
    """
    Create a new daily tracker Excel file from template
    
    Args:
        instance: Instance name (prod, dev, testing)
        username: Username to include in filename
        tracker_name: Name of the tracker
        tracker_type: Type of tracker ('50K', '1L', 'No Reinvest')
        investment: Investment amount
        scheme_period: Scheme period in days
        start_date: Start date (date object or string)
        per_day_payment: Per day payment amount (optional, will use defaults if not provided)
    
    Returns:
        Filename of created tracker
    """
    # Generate unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()
    filename = f"tracker_{safe_username}_{timestamp}.xlsx"
    
    # Get template path
    template_path = Path("daily-trackers/template/DailyTrackerTemplate.xlsx")
    
    # Get instance tracker directory
    tracker_dir = Path(get_tracker_directory(instance))
    tracker_path = tracker_dir / filename
    
    # Copy template to new location
    shutil.copy(str(template_path), str(tracker_path))
    
    # Load workbook and get the sheet
    wb = openpyxl.load_workbook(str(tracker_path))
    
    # Get sheet configuration
    sheet_name = TRACKER_TYPES.get(tracker_type)
    if not sheet_name:
        raise ValueError(f"Invalid tracker type: {tracker_type}")
    
    config = SHEET_CONFIGS.get(tracker_type)
    if not config:
        raise ValueError(f"No configuration found for tracker type: {tracker_type}")
    
    # Select the appropriate sheet
    ws = wb[sheet_name]
    
    # Fill in the parameters
    ws[config['tracker_name_cell']] = tracker_name
    ws[config['investment_cell']] = float(investment)
    ws[config['scheme_period_cell']] = f"{scheme_period} (Days)"
    
    # Handle start_date
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    ws[config['start_date_cell']] = start_date
    
    # Set per day payment - use provided value or defaults based on tracker type
    if per_day_payment is None:
        # Use default values if not provided
        if tracker_type == '50K':
            per_day_payment = 500
        elif tracker_type == '1L':
            per_day_payment = 1000
        else:  # No Reinvest
            per_day_payment = 3000
    
    ws[config['per_day_payment_cell']] = float(per_day_payment)
    
    # Remove other sheets (keep only the selected tracker type)
    for sheet in wb.sheetnames:
        if sheet != sheet_name:
            del wb[sheet]
    
    # Save the workbook
    wb.save(str(tracker_path))
    
    return filename


def get_tracker_data(instance, filename):
    """
    Read all data from a tracker Excel file
    
    Args:
        instance: Instance name
        filename: Filename of the tracker
    
    Returns:
        Dictionary containing tracker parameters and data rows
    """
    tracker_dir = Path(get_tracker_directory(instance))
    tracker_path = tracker_dir / filename
    
    if not tracker_path.exists():
        raise FileNotFoundError(f"Tracker file not found: {filename}")
    
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True)
    ws = wb.active  # Since we only keep one sheet
    
    sheet_name = ws.title
    
    # Determine tracker type from sheet name
    tracker_type = None
    for key, value in TRACKER_TYPES.items():
        if value == sheet_name:
            tracker_type = key
            break
    
    if not tracker_type:
        raise ValueError(f"Unknown sheet name: {sheet_name}")
    
    config = SHEET_CONFIGS[tracker_type]
    
    # Read parameters
    parameters = {
        'tracker_name': ws[config['tracker_name_cell']].value,
        'investment': ws[config['investment_cell']].value,
        'scheme_period': ws[config['scheme_period_cell']].value,
        'start_date': ws[config['start_date_cell']].value,
        'per_day_payment': ws[config['per_day_payment_cell']].value,
        'tracker_type': tracker_type
    }
    
    # Read summary section (if available) - just read from the cells
    # For now, we'll calculate on the fly from data
    
    # Read data rows
    data_rows = []
    data_start = config['data_start_row']
    columns = config['columns']
    
    # Read up to 400 rows (more than enough for most schemes)
    for row_num in range(data_start, data_start + 400):
        day_cell = f"{columns['day']}{row_num}"
        day_value = ws[day_cell].value
        
        # Skip if day is empty
        if day_value is None or day_value == '':
            continue
        
        row_data = {
            'row_num': row_num,
            'day': day_value
        }
        
        # Read all columns based on tracker type
        for col_name, col_letter in columns.items():
            if col_name == 'day':
                continue
            cell = f"{col_letter}{row_num}"
            value = ws[cell].value
            
            # Handle date formatting
            if col_name == 'date' and isinstance(value, datetime):
                value = value.date()
            
            row_data[col_name] = value
        
        data_rows.append(row_data)
    
    return {
        'parameters': parameters,
        'data': data_rows,
        'tracker_type': tracker_type,
        'sheet_name': sheet_name
    }


def update_tracker_entry(instance, filename, day, entry_data):
    """
    Update a specific day's entry in the tracker
    
    Args:
        instance: Instance name
        filename: Filename of the tracker
        day: Day number to update
        entry_data: Dictionary with column names and values to update
    
    Returns:
        True if successful
    """
    tracker_dir = Path(get_tracker_directory(instance))
    tracker_path = tracker_dir / filename
    
    if not tracker_path.exists():
        raise FileNotFoundError(f"Tracker file not found: {filename}")
    
    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb.active
    
    sheet_name = ws.title
    
    # Determine tracker type from sheet name
    tracker_type = None
    for key, value in TRACKER_TYPES.items():
        if value == sheet_name:
            tracker_type = key
            break
    
    if not tracker_type:
        raise ValueError(f"Unknown sheet name: {sheet_name}")
    
    config = SHEET_CONFIGS[tracker_type]
    columns = config['columns']
    data_start = config['data_start_row']
    
    # Find the row for this day
    row_num = None
    for check_row in range(data_start, data_start + 400):
        day_cell = f"{columns['day']}{check_row}"
        if ws[day_cell].value == day:
            row_num = check_row
            break
    
    if row_num is None:
        raise ValueError(f"Day {day} not found in tracker")
    
    # Update the specified columns
    for col_name, value in entry_data.items():
        if col_name in columns:
            col_letter = columns[col_name]
            cell = f"{col_letter}{row_num}"
            
            # Handle date conversion
            if col_name == 'date':
                if isinstance(value, str):
                    value = datetime.strptime(value, '%Y-%m-%d').date()
            
            # Convert Decimal to float for Excel
            if isinstance(value, Decimal):
                value = float(value)
            
            ws[cell] = value
    
    # Auto-calculate cumulative if daily_payments is provided but cumulative is not
    if 'daily_payments' in entry_data and 'cumulative' not in entry_data:
        # Calculate cumulative as sum of all daily_payments from start to this day
        cumulative_sum = 0
        for calc_row in range(data_start, row_num + 1):
            payment_cell = f"{columns['daily_payments']}{calc_row}"
            payment_value = ws[payment_cell].value
            if payment_value is not None:
                try:
                    cumulative_sum += float(payment_value)
                except:
                    pass
        
        # Set the cumulative value
        cumulative_cell = f"{columns['cumulative']}{row_num}"
        ws[cumulative_cell] = cumulative_sum
    
    # Save the workbook
    wb.save(str(tracker_path))
    
    return True


def get_tracker_summary(instance, filename):
    """
    Get summary information from a tracker
    
    Args:
        instance: Instance name
        filename: Filename of the tracker
    
    Returns:
        Dictionary containing summary information
    """
    data = get_tracker_data(instance, filename)
    parameters = data['parameters']
    rows = data['data']
    
    # Calculate summary
    total_days = len([r for r in rows if r.get('daily_payments')])
    total_payments = sum(float(r.get('daily_payments', 0) or 0) for r in rows)
    
    # Get latest balance/cumulative and highest day number
    latest_row = None
    highest_day = 0
    for row in reversed(rows):
        if row.get('daily_payments') is not None:
            if latest_row is None:
                latest_row = row
            # Track the highest day number with a payment
            day_num = row.get('day', 0)
            if isinstance(day_num, (int, float)) and day_num > highest_day:
                highest_day = int(day_num)
    
    balance = 0
    cumulative = 0
    pending = 0
    
    if latest_row:
        if data['tracker_type'] in ['50K', '1L']:
            balance = latest_row.get('balance', 0) or 0
        cumulative = latest_row.get('cumulative', 0) or 0
    
    # Calculate pending based on highest day number, not count of days
    # If Day 2 has payment, expected should be 2 * per_day, not 1 * per_day
    expected_total = highest_day * float(parameters.get('per_day_payment', 0) or 0)
    pending = expected_total - total_payments
    
    return {
        'total_days': total_days,
        'total_payments': total_payments,
        'balance': balance,
        'cumulative': cumulative,
        'pending': pending,
        'expected_total': expected_total,
        'highest_day': highest_day,
        'parameters': parameters
    }

