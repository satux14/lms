"""
Multi-Instance Lending Management System
======================================

This is a simplified multi-instance application that handles prod, dev, and testing instances
in a single Flask app with URL-based routing.

URL Structure:
- /prod/... - Production instance
- /dev/... - Development instance  
- /testing/... - Testing instance
- /... - Default to production instance

Author: Lending Management System
Version: 1.0.1
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, g, abort, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_babel import gettext, lazy_gettext
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from sqlalchemy import create_engine, or_
from sqlalchemy.orm import sessionmaker
import os
import sys
import uuid
import secrets
import smtplib
import hmac
import hashlib
import json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

# Google Pay UPI uses Payment Request API (browser native) - no server-side SDK needed

# Add daily-trackers to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'daily-trackers'))
from tracker_manager import (
    create_tracker_file, get_tracker_data, update_tracker_entry, 
    get_tracker_summary, TRACKER_TYPES, get_tracker_directory
)

# Import i18n module (can be disabled by commenting out)
from i18n_config import init_i18n, get_supported_languages, get_current_language

# Constants
DAYS_PER_YEAR = 360  # Interest calculation based on 360 days per year

# Instance configuration
DEFAULT_INSTANCE = 'prod'
VALID_INSTANCES = ['prod', 'dev', 'testing']

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Google Pay UPI configuration
app.config['GOOGLE_PAY_MERCHANT_VPA'] = os.environ.get('GOOGLE_PAY_MERCHANT_VPA', '')
app.config['GOOGLE_PAY_MERCHANT_NAME'] = os.environ.get('GOOGLE_PAY_MERCHANT_NAME', 'The SRS Consulting')
app.config['GOOGLE_PAY_MERCHANT_CODE'] = os.environ.get('GOOGLE_PAY_MERCHANT_CODE', '0000')  # MCC code
app.config['GOOGLE_PAY_CALLBACK_URL'] = os.environ.get('GOOGLE_PAY_CALLBACK_URL', '')

# Initialize extensions without database URI first
db = SQLAlchemy()
login_manager = LoginManager()
login_manager.login_view = 'login_redirect'

# Initialize internationalization (i18n) - can be disabled by commenting out
init_i18n(app)

# We'll initialize with app later after setting the correct database URI

@login_manager.user_loader
def load_user(user_id):
    try:
        # Get current instance from g or default to prod
        instance = getattr(g, 'current_instance', 'prod')
        
        # Ensure database manager is initialized
        if not db_manager.initialized:
            db_manager.initialize_all_databases()
        
        # Use session.get() instead of query.get() for SQLAlchemy 2.0 compatibility
        session = db_manager.get_session_for_instance(instance)
        user = session.get(User, int(user_id))
        return user
    except Exception as e:
        # If there's any error, return None (user not found)
        return None

@app.context_processor
def inject_cashback_balance():
    """Inject cashback balance into all template contexts"""
    try:
        if current_user.is_authenticated:
            instance = getattr(g, 'current_instance', 'prod')
            balance = get_user_cashback_balance(current_user.id, instance)
            return {'cashback_balance': balance, 'current_instance': instance}
    except Exception:
        pass
    return {'cashback_balance': Decimal('0'), 'current_instance': getattr(g, 'current_instance', 'prod')}

# Database Models (same as original)
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_moderator = db.Column(db.Boolean, default=False)
    language_preference = db.Column(db.String(5), default='en')  # Language preference (en/ta)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    reset_token = db.Column(db.String(100), nullable=True)
    reset_token_expires = db.Column(db.DateTime, nullable=True)

class InterestRate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rate = db.Column(db.Numeric(10, 4), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)


# Association tables for moderator access
moderator_loans = db.Table('moderator_loans',
    db.Column('moderator_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('loan_id', db.Integer, db.ForeignKey('loan.id'), primary_key=True),
    db.Column('assigned_at', db.DateTime, default=datetime.utcnow)
)

moderator_trackers = db.Table('moderator_trackers',
    db.Column('moderator_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('tracker_id', db.Integer, db.ForeignKey('daily_tracker.id'), primary_key=True),
    db.Column('assigned_at', db.DateTime, default=datetime.utcnow)
)


class Loan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    loan_name = db.Column(db.String(100), nullable=False)
    principal_amount = db.Column(db.Numeric(15, 2), nullable=False)
    remaining_principal = db.Column(db.Numeric(15, 2), nullable=False)
    interest_rate = db.Column(db.Numeric(5, 4), nullable=False)
    payment_frequency = db.Column(db.String(20), nullable=False)
    loan_type = db.Column(db.String(20), nullable=False, default='regular')
    admin_notes = db.Column(db.Text, nullable=True)
    customer_notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    status = db.Column(db.String(20), default='active')  # 'active', 'closed', 'deleted'
    
    # Relationships
    customer = db.relationship('User', backref='loans')
    payments = db.relationship('Payment', primaryjoin='Loan.id == Payment.loan_id', backref='loan', lazy=True, order_by='Payment.payment_date.desc()')
    assigned_moderators = db.relationship('User', secondary=moderator_loans, 
                                         backref=db.backref('assigned_loans', lazy='dynamic'))

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    amount = db.Column(db.Numeric(15, 2), nullable=False)
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)
    payment_type = db.Column(db.String(20), nullable=False)
    interest_amount = db.Column(db.Numeric(15, 2), default=0)
    principal_amount = db.Column(db.Numeric(15, 2), default=0)
    transaction_id = db.Column(db.String(100), nullable=True)
    payment_method = db.Column(db.String(20), nullable=True)
    proof_filename = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(20), default='pending')
    # Razorpay fields
    razorpay_order_id = db.Column(db.String(100), nullable=True)
    razorpay_payment_id = db.Column(db.String(100), nullable=True)
    razorpay_signature = db.Column(db.String(255), nullable=True)
    payment_initiated_at = db.Column(db.DateTime, nullable=True)
    # Loan splitting fields
    split_loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=True)  # For split loan assignment
    original_principal_amount = db.Column(db.Numeric(15, 2), nullable=True)  # Principal amount at time of payment

class LoanSplit(db.Model):
    """Model to track loan splits - when a loan is split into multiple parts"""
    id = db.Column(db.Integer, primary_key=True)
    original_loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    split_loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    split_principal_amount = db.Column(db.Numeric(15, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Relationships
    original_loan = db.relationship('Loan', foreign_keys=[original_loan_id], backref='splits_from')
    split_loan = db.relationship('Loan', foreign_keys=[split_loan_id], backref='split_from')
    creator = db.relationship('User', backref='loan_splits_created')

class PendingInterest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    amount = db.Column(db.Numeric(15, 2), nullable=False)
    due_date = db.Column(db.Date, nullable=False)
    month_year = db.Column(db.String(7), nullable=False)
    is_paid = db.Column(db.Boolean, default=False)

class DailyTracker(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    tracker_name = db.Column(db.String(200), nullable=False)
    tracker_type = db.Column(db.String(50), nullable=False)  # '50K', '1L', 'No Reinvest'
    investment = db.Column(db.Numeric(15, 2), nullable=False)
    scheme_period = db.Column(db.Integer, nullable=False)  # in days
    per_day_payment = db.Column(db.Numeric(15, 2), nullable=False)  # daily payment amount
    start_date = db.Column(db.Date, nullable=False)
    filename = db.Column(db.String(255), nullable=False)  # unique filename for the Excel file
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)  # Admin-level active/deleted
    is_closed_by_user = db.Column(db.Boolean, default=False)  # User has closed/hidden tracker
    
    # Relationships
    user = db.relationship('User', backref='daily_trackers')
    assigned_moderators = db.relationship('User', secondary=moderator_trackers,
                                         backref=db.backref('assigned_trackers', lazy='dynamic'))

class CashbackTransaction(db.Model):
    """Model to track all cashback point transactions"""
    id = db.Column(db.Integer, primary_key=True)
    from_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # NULL for system/admin grants
    to_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    points = db.Column(db.Numeric(15, 2), nullable=False)
    transaction_type = db.Column(db.String(30), nullable=False)  # 'transfer', 'loan_interest_auto', 'loan_interest_manual', 'tracker_entry', 'unconditional', 'system'
    related_loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=True)
    related_payment_id = db.Column(db.Integer, db.ForeignKey('payment.id'), nullable=True)
    related_tracker_id = db.Column(db.Integer, db.ForeignKey('daily_tracker.id'), nullable=True)
    related_tracker_entry_day = db.Column(db.Integer, nullable=True)
    notes = db.Column(db.Text, nullable=True)  # Admin/internal notes
    user_notes = db.Column(db.Text, nullable=True)  # User-visible notes/reason
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    from_user = db.relationship('User', foreign_keys=[from_user_id], backref='cashback_sent')
    to_user = db.relationship('User', foreign_keys=[to_user_id], backref='cashback_received')
    related_loan = db.relationship('Loan', backref='cashback_transactions')
    related_payment = db.relationship('Payment', backref='cashback_transactions')
    related_tracker = db.relationship('DailyTracker', backref='cashback_transactions')
    created_by = db.relationship('User', foreign_keys=[created_by_user_id], backref='cashback_created')

class LoanCashbackConfig(db.Model):
    """Model to store per-loan cashback configuration"""
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cashback_type = db.Column(db.String(20), nullable=False)  # 'percentage' or 'fixed'
    cashback_value = db.Column(db.Numeric(15, 4), nullable=False)  # percentage as decimal (0.05 for 5%) or fixed amount
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationships
    loan = db.relationship('Loan', backref='cashback_configs')
    user = db.relationship('User', backref='loan_cashback_configs')

class TrackerEntry(db.Model):
    """Model to track individual tracker entries with approval status"""
    id = db.Column(db.Integer, primary_key=True)
    tracker_id = db.Column(db.Integer, db.ForeignKey('daily_tracker.id'), nullable=False)
    day = db.Column(db.Integer, nullable=False)
    entry_data = db.Column(db.Text, nullable=False)  # JSON string of entry data
    status = db.Column(db.String(20), nullable=False, default='pending')  # 'pending', 'verified', 'rejected'
    submitted_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    verified_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    verified_at = db.Column(db.DateTime, nullable=True)
    rejection_reason = db.Column(db.Text, nullable=True)
    
    tracker = db.relationship('DailyTracker', backref='tracker_entries')
    submitted_by = db.relationship('User', foreign_keys=[submitted_by_user_id], backref='submitted_tracker_entries')
    verified_by = db.relationship('User', foreign_keys=[verified_by_user_id], backref='verified_tracker_entries')

class TrackerCashbackConfig(db.Model):
    """Model to store per-tracker cashback configuration"""
    id = db.Column(db.Integer, primary_key=True)
    tracker_id = db.Column(db.Integer, db.ForeignKey('daily_tracker.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cashback_type = db.Column(db.String(20), nullable=False)  # 'percentage' or 'fixed'
    cashback_value = db.Column(db.Numeric(15, 4), nullable=False)  # percentage as decimal (0.05 for 5%) or fixed amount
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    tracker = db.relationship('DailyTracker', backref='cashback_configs')
    user = db.relationship('User', backref='tracker_cashback_configs')

class UserPaymentMethod(db.Model):
    """Model to store user's payment method details for cashback redemption"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    payment_type = db.Column(db.String(50), nullable=False)  # 'cash', 'gpay', 'upi', 'account', 'amazon_gift_card'
    account_name = db.Column(db.String(200), nullable=True)  # Account holder name
    account_number = db.Column(db.String(100), nullable=True)  # Account number, UPI ID, GPay ID, etc.
    ifsc_code = db.Column(db.String(20), nullable=True)  # For bank accounts
    bank_name = db.Column(db.String(200), nullable=True)  # Bank name
    upi_id = db.Column(db.String(200), nullable=True)  # UPI ID
    gpay_id = db.Column(db.String(200), nullable=True)  # GPay ID/Phone number
    phone_number = db.Column(db.String(20), nullable=True)  # Phone number for GPay/cash
    address = db.Column(db.Text, nullable=True)  # Address for cash delivery
    is_default = db.Column(db.Boolean, default=False)  # Default payment method
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref='payment_methods')

class CashbackRedemption(db.Model):
    """Model to track cashback redemption requests"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount = db.Column(db.Numeric(15, 2), nullable=False)  # Amount to redeem
    redemption_type = db.Column(db.String(50), nullable=False)  # 'cash', 'gpay', 'upi', 'account', 'amazon_gift_card'
    payment_method_id = db.Column(db.Integer, db.ForeignKey('user_payment_method.id'), nullable=True)  # Saved payment method
    # Payment details (can be from saved method or manually entered)
    account_name = db.Column(db.String(200), nullable=True)
    account_number = db.Column(db.String(100), nullable=True)
    ifsc_code = db.Column(db.String(20), nullable=True)
    bank_name = db.Column(db.String(200), nullable=True)
    upi_id = db.Column(db.String(200), nullable=True)
    gpay_id = db.Column(db.String(200), nullable=True)
    phone_number = db.Column(db.String(20), nullable=True)
    address = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), nullable=False, default='pending')  # 'pending', 'completed', 'cancelled'
    admin_notes = db.Column(db.Text, nullable=True)  # Admin notes when processing
    processed_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Admin who processed
    processed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Create cashback transaction for redemption (deduct points)
    redemption_transaction_id = db.Column(db.Integer, db.ForeignKey('cashback_transaction.id'), nullable=True)
    
    user = db.relationship('User', foreign_keys=[user_id], backref='cashback_redemptions')
    payment_method = db.relationship('UserPaymentMethod', backref='redemptions')
    processed_by = db.relationship('User', foreign_keys=[processed_by_user_id], backref='processed_redemptions')
    redemption_transaction = db.relationship('CashbackTransaction', foreign_keys=[redemption_transaction_id], backref='redemption')

class NotificationPreference(db.Model):
    """Model to store user notification preferences"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True)
    channel = db.Column(db.String(20), nullable=False, default='email')  # 'email', 'sms', 'slack', etc.
    enabled = db.Column(db.Boolean, nullable=False, default=True)  # Master switch for this channel
    preferences = db.Column(db.JSON, nullable=True)  # Channel-specific preferences as JSON
    # Example preferences for email: {'payment_approvals': True, 'tracker_approvals': True, 'payment_status': False, 'tracker_status': False}
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref='notification_preferences')

# Instance management

def get_current_instance():
    """Get current instance from URL path"""
    try:
        path_parts = request.path.strip('/').split('/')
        if path_parts and path_parts[0] in VALID_INSTANCES:
            return path_parts[0]
        return DEFAULT_INSTANCE
    except:
        return DEFAULT_INSTANCE

def get_database_uri(instance=None):
    """Get database URI for specific instance"""
    if instance is None:
        instance = get_current_instance()
    
    if instance not in VALID_INSTANCES:
        instance = DEFAULT_INSTANCE
    
    # Create instances directory (plural) - consolidated location
    instances_dir = Path("instances")
    instance_dir = instances_dir / instance
    instance_dir.mkdir(parents=True, exist_ok=True)
    
    # Create database directory
    db_dir = instance_dir / "database"
    db_dir.mkdir(exist_ok=True)
    
    db_path = db_dir / f"lending_app_{instance}.db"
    return f"sqlite:///{db_path.absolute()}"

def get_uploads_folder(instance=None):
    """Get uploads folder for specific instance"""
    if instance is None:
        instance = get_current_instance()
    
    if instance not in VALID_INSTANCES:
        instance = DEFAULT_INSTANCE
    
    instances_dir = Path("instances")
    instance_dir = instances_dir / instance
    uploads_dir = instance_dir / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    
    return str(uploads_dir)

def configure_app_for_instance(instance):
    """Configure app for specific instance"""
    app.config['SQLALCHEMY_DATABASE_URI'] = get_database_uri(instance)
    app.config['UPLOAD_FOLDER'] = get_uploads_folder(instance)
    app.config['INSTANCE_NAME'] = instance

# Custom database manager for proper instance isolation
class DatabaseManager:
    def __init__(self):
        self.databases = {}
        self.engines = {}
        self.sessions = {}
        self.initialized = False
    
    def initialize_all_databases(self):
        """Initialize all database connections for all instances"""
        if self.initialized:
            return
        
        print("Initializing database connections for all instances...")
        
        for instance in VALID_INSTANCES:
            print(f"  Initializing {instance} database...")
            
            # Configure the database URI
            instance_uri = get_database_uri(instance)
            
            # Create engine for this instance
            engine = create_engine(instance_uri)
            self.engines[instance] = engine
            
            # Create session for this instance
            session = sessionmaker(bind=engine)()
            self.sessions[instance] = session
            
            # Create tables using the engine
            with engine.connect() as conn:
                # Import the models to ensure they're registered
                from app_multi import User, Loan, Payment
                
                # Create all tables
                db.metadata.create_all(engine)
        
        self.initialized = True
        print("All database connections initialized successfully!")
    
    
    def get_engine_for_instance(self, instance):
        """Get engine for specific instance"""
        if not self.initialized:
            self.initialize_all_databases()
        
        if instance not in self.engines:
            raise ValueError(f"Engine for instance '{instance}' not found")
        
        return self.engines[instance]
    
    def get_session_for_instance(self, instance):
        """Get session for specific instance"""
        if not self.initialized:
            self.initialize_all_databases()
        
        if instance not in self.sessions:
            raise ValueError(f"Session for instance '{instance}' not found")
        
        return self.sessions[instance]
    
    
    def get_query_for_instance(self, instance, model_class):
        """Get query object for specific instance and model"""
        if not self.initialized:
            self.initialize_all_databases()
        
        if instance not in self.sessions:
            raise ValueError(f"Session for instance '{instance}' not found")
        
        session = self.sessions[instance]
        return session.query(model_class)
    
    def add_to_instance(self, instance, obj):
        """Add object to specific instance database"""
        if not self.initialized:
            self.initialize_all_databases()
        
        if instance not in self.sessions:
            raise ValueError(f"Session for instance '{instance}' not found")
        
        session = self.sessions[instance]
        session.add(obj)
        session.commit()
        return obj

# Global database manager
db_manager = DatabaseManager()

# Initialize logging and metrics for each instance
from lms_logging import init_logging, get_logging_manager
from lms_metrics import init_metrics, get_metrics_manager

def initialize_logging_and_metrics():
    """Initialize logging and metrics for all instances"""
    for instance in VALID_INSTANCES:
        # Get engine for this instance
        engine = db_manager.get_engine_for_instance(instance)
        
        # Initialize logging (same DB, different tables)
        init_logging(instance_name=instance, db_engine=engine)
        
        # Initialize metrics (same DB, different tables)
        init_metrics(instance_name=instance, db_engine=engine)
    
    print("Logging and metrics initialized for all instances!")

# Custom database initialization function
def init_database_for_instance(instance):
    """Initialize database for specific instance"""
    instance_uri = get_database_uri(instance)
    app.config['SQLALCHEMY_DATABASE_URI'] = instance_uri
    
    # Create all tables
    with app.app_context():
        db.create_all()

# Initialize app
def init_app():
    """Initialize the application"""
    # Configure for default instance first
    configure_app_for_instance(DEFAULT_INSTANCE)
    
    # Initialize database and login manager with the configured app
    db.init_app(app)
    login_manager.init_app(app)
    
    # Initialize all database connections
    db_manager.initialize_all_databases()
    
    # Initialize logging and metrics for all instances
    initialize_logging_and_metrics()
    
    # Register payment routes
    from app_payments import register_payment_routes, process_payment as payment_process_payment
    register_payment_routes(
        flask_app=app,
        flask_db=db,
        valid_instances=VALID_INSTANCES,
        payment_model=Payment,
        loan_model=Loan,
        payment_query_func=get_payment_query,
        loan_query_func=get_loan_query,
        add_instance_func=add_to_current_instance,
        commit_instance_func=commit_current_instance,
        verify_payment_func=verify_payment,
        calc_accumulated_func=calculate_accumulated_interest,
        calc_daily_func=calculate_daily_interest,
        calc_monthly_func=calculate_monthly_interest
    )
    # Make process_payment available for admin routes
    global process_payment
    process_payment = payment_process_payment
    
    # Register tracker routes
    from app_trackers import register_tracker_routes
    register_tracker_routes(
        flask_app=app,
        flask_db=db,
        valid_instances=VALID_INSTANCES,
        default_instance=DEFAULT_INSTANCE,
        tracker_model=DailyTracker,
        tracker_entry_model=TrackerEntry,
        tracker_cashback_config_model=TrackerCashbackConfig,
        user_model=User,
        cashback_transaction_model=CashbackTransaction,
        tracker_query_func=get_daily_tracker_query,
        tracker_entry_query_func=get_tracker_entry_query,
        tracker_cashback_config_query_func=get_tracker_cashback_config_query,
        user_query_func=get_user_query,
        cashback_transaction_query_func=get_cashback_transaction_query,
        add_instance_func=add_to_current_instance,
        commit_instance_func=commit_current_instance,
        get_current_instance_func=get_current_instance_from_g,
        db_manager_instance=db_manager,
        validate_username_exists_helper=validate_username_exists,
        get_user_cashback_balance_helper=get_user_cashback_balance,
        create_tracker_file_func=create_tracker_file,
        get_tracker_data_func=get_tracker_data,
        update_tracker_entry_func=update_tracker_entry,
        get_tracker_summary_func=get_tracker_summary,
        tracker_types=TRACKER_TYPES,
        get_tracker_directory_func=get_tracker_directory
    )
    
    # Register loan routes
    from app_loans import register_loan_routes
    register_loan_routes(
        flask_app=app,
        flask_db=db,
        valid_instances=VALID_INSTANCES,
        default_instance=DEFAULT_INSTANCE,
        loan_model=Loan,
        payment_model=Payment,
        loan_split_model=LoanSplit,
        loan_cashback_config_model=LoanCashbackConfig,
        user_model=User,
        cashback_transaction_model=CashbackTransaction,
        loan_query_func=get_loan_query,
        loan_split_query_func=get_loan_split_query,
        loan_cashback_config_query_func=get_loan_cashback_config_query,
        payment_query_func=get_payment_query,
        user_query_func=get_user_query,
        cashback_transaction_query_func=get_cashback_transaction_query,
        add_instance_func=add_to_current_instance,
        commit_instance_func=commit_current_instance,
        get_current_instance_func=get_current_instance_from_g,
        db_manager_instance=db_manager,
        validate_username_exists_helper=validate_username_exists,
        get_payment_cashback_total_helper=get_payment_cashback_total,
        calculate_daily_interest_func=calculate_daily_interest,
        calculate_monthly_interest_func=calculate_monthly_interest,
        calculate_accumulated_interest_func=calculate_accumulated_interest,
        calculate_interest_for_period_func=calculate_interest_for_period,
        days_per_year=DAYS_PER_YEAR,
        get_logging_manager_func=get_logging_manager,
        get_metrics_manager_func=get_metrics_manager
    )
    
    # Register moderator routes
    from app_moderator import register_moderator_routes
    register_moderator_routes(
        flask_app=app,
        flask_db=db,
        valid_instances=VALID_INSTANCES,
        default_instance=DEFAULT_INSTANCE,
        payment_model=Payment,
        loan_model=Loan,
        tracker_model=DailyTracker,
        payment_query_func=get_payment_query,
        loan_query_func=get_loan_query,
        user_query_func=get_user_query,
        tracker_query_func=get_daily_tracker_query,
        add_instance_func=add_to_current_instance,
        commit_instance_func=commit_current_instance,
        calc_accumulated_func=calculate_accumulated_interest,
        calc_daily_interest_func=calculate_daily_interest,
        calc_monthly_interest_func=calculate_monthly_interest,
        process_payment_func=process_payment,
        get_tracker_data_func=get_tracker_data,
        get_tracker_summary_func=get_tracker_summary,
        update_tracker_entry_func=update_tracker_entry
    )
    
    # Register cashback routes
    from app_cashback import register_cashback_routes
    register_cashback_routes(
        flask_app=app,
        flask_db=db,
        valid_instances=VALID_INSTANCES,
        default_instance=DEFAULT_INSTANCE,
        user_model=User,
        loan_model=Loan,
        tracker_model=DailyTracker,
        payment_model=Payment,
        cashback_transaction_model=CashbackTransaction,
        loan_cashback_config_model=LoanCashbackConfig,
        tracker_cashback_config_model=TrackerCashbackConfig,
        user_payment_method_model=UserPaymentMethod,
        cashback_redemption_model=CashbackRedemption,
        user_query_func=get_user_query,
        loan_query_func=get_loan_query,
        tracker_query_func=get_daily_tracker_query,
        payment_query_func=get_payment_query,
        cashback_transaction_query_func=get_cashback_transaction_query,
        loan_cashback_config_query_func=get_loan_cashback_config_query,
        tracker_cashback_config_query_func=get_tracker_cashback_config_query,
        user_payment_method_query_func=get_user_payment_method_query,
        cashback_redemption_query_func=get_cashback_redemption_query,
        add_instance_func=add_to_current_instance,
        commit_instance_func=commit_current_instance,
        get_current_instance_func=get_current_instance_from_g,
        db_manager_instance=db_manager,
        get_user_cashback_balance_helper=get_user_cashback_balance,
        get_loan_cashback_total_helper=lambda loan_id, instance: __import__('app_loans', fromlist=['get_loan_cashback_total']).get_loan_cashback_total(loan_id, instance),
        get_tracker_cashback_total_helper=lambda tracker_id, instance: __import__('app_trackers', fromlist=['get_tracker_cashback_total']).get_tracker_cashback_total(tracker_id, instance),
        get_tracker_day_cashback_helper=lambda tracker_id, day, instance: __import__('app_trackers', fromlist=['get_tracker_day_cashback']).get_tracker_day_cashback(tracker_id, day, instance),
        get_payment_cashback_total_helper=get_payment_cashback_total,
        process_loan_cashback_helper=lambda loan, payment, instance, user_id: __import__('app_loans', fromlist=['process_loan_cashback']).process_loan_cashback(loan, payment, instance, user_id),
        validate_username_exists_helper=validate_username_exists
    )
    
    # Register backup routes
    from app_backup import register_backup_routes
    register_backup_routes(
        flask_app=app,
        valid_instances=VALID_INSTANCES
    )
    
    # Create default data for all instances
    with app.app_context():
        for instance in VALID_INSTANCES:
            # Set the instance in g for create_default_data
            g.current_instance = instance
            create_default_data(instance)
    
    return app

def create_default_data(instance):
    """Create default data for instance"""
    # Create default admin user if it doesn't exist
    User_query = db_manager.get_query_for_instance(instance, User)
    if not User_query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            email=f'admin@{instance}.lendingapp.com',
            password_hash=generate_password_hash('admin123'),
            is_admin=True
        )
        db_manager.add_to_instance(instance, admin)
        
        # Create default interest rate
        default_rate = InterestRate(rate=Decimal('0.21'))  # 21%
        db_manager.add_to_instance(instance, default_rate)
        
        print(f"Default admin user created for {instance}: username='admin', password='admin123'")

# Helper functions
def get_current_instance_from_g():
    """Get current instance from Flask g object"""
    return getattr(g, 'current_instance', 'prod')

def get_user_query():
    """Get User query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, User)

def get_loan_query():
    """Get Loan query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, Loan)

def get_payment_query():
    """Get Payment query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, Payment)

def get_interest_rate_query():
    """Get InterestRate query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, InterestRate)

def get_daily_tracker_query():
    """Get DailyTracker query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, DailyTracker)

def get_cashback_transaction_query():
    """Get CashbackTransaction query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, CashbackTransaction)

def get_loan_split_query():
    """Get LoanSplit query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, LoanSplit)

def get_loan_cashback_config_query():
    """Get LoanCashbackConfig query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, LoanCashbackConfig)

def get_tracker_entry_query():
    """Get TrackerEntry query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, TrackerEntry)

def get_user_payment_method_query():
    """Get UserPaymentMethod query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, UserPaymentMethod)

def get_cashback_redemption_query():
    """Get CashbackRedemption query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, CashbackRedemption)

def get_tracker_cashback_config_query():
    """Get TrackerCashbackConfig query for current instance"""
    instance = get_current_instance_from_g()
    return db_manager.get_query_for_instance(instance, TrackerCashbackConfig)

def get_payment_cashback_total(payment_id, instance_name):
    """Get total cashback given for a specific payment"""
    session = db_manager.get_session_for_instance(instance_name)
    result = session.query(
        db.func.sum(CashbackTransaction.points)
    ).filter_by(
        related_payment_id=payment_id
    ).scalar()
    return result or Decimal('0')

def add_to_current_instance(obj):
    """Add object to current instance database"""
    instance = get_current_instance_from_g()
    return db_manager.add_to_instance(instance, obj)

def commit_current_instance():
    """Commit current instance database"""
    instance = get_current_instance_from_g()
    session = db_manager.get_session_for_instance(instance)
    session.commit()

# Cashback helper functions
def get_user_cashback_balance(user_id, instance_name):
    """Calculate user's cashback balance from transactions"""
    try:
        # Get session for the instance
        session = db_manager.get_session_for_instance(instance_name)
        
        # Check if this is a system user (admin)
        # System users should not accumulate balance from deduction/redemption transactions
        user = session.query(User).filter_by(id=user_id).first()
        is_system_user = user and user.is_admin
        
        # Sum of all points received
        received_query = session.query(
            db.func.sum(CashbackTransaction.points)
        ).filter_by(to_user_id=user_id)
        
        # Exclude deduction and redemption transactions for system users
        # (these should not increase system user balance - they're just accounting entries)
        if is_system_user:
            received_query = received_query.filter(
                ~CashbackTransaction.transaction_type.in_(['deduction', 'redemption'])
            )
        
        received = received_query.scalar() or Decimal('0')
        
        # Sum of all points sent
        sent = session.query(
            db.func.sum(CashbackTransaction.points)
        ).filter_by(from_user_id=user_id).scalar() or Decimal('0')
        
        balance = received - sent
        return balance
    except Exception as e:
        print(f"Error calculating cashback balance: {e}")
        return Decimal('0')

def validate_username_exists(username, instance_name):
    """Check if username exists and return user object or None"""
    try:
        session = db_manager.get_session_for_instance(instance_name)
        user = session.query(User).filter_by(username=username).first()
        return user
    except Exception as e:
        print(f"Error validating username: {e}")
        return None

# Tracker helper functions moved to app_trackers.py

# Loan helper functions moved to app_loans.py

def calculate_daily_interest(principal, annual_rate):
    """Calculate daily interest amount"""
    try:
        daily_rate = annual_rate / DAYS_PER_YEAR
        return Decimal(str(principal)) * Decimal(str(daily_rate))
    except (InvalidOperation, TypeError):
        return Decimal('0')

def calculate_monthly_interest(principal, annual_rate):
    """Calculate monthly interest amount"""
    try:
        monthly_rate = annual_rate / 12
        return Decimal(str(principal)) * Decimal(str(monthly_rate))
    except (InvalidOperation, TypeError):
        return Decimal('0')

def calculate_interest_for_period(principal, annual_rate, start_date, end_date):
    """Calculate interest for a specific time period"""
    try:
        days = (end_date - start_date).days
        if days <= 0:
            return Decimal('0')
        
        daily_rate = annual_rate / DAYS_PER_YEAR
        return Decimal(str(principal)) * Decimal(str(daily_rate)) * Decimal(str(days))
    except (InvalidOperation, TypeError):
        return Decimal('0')

def calculate_accumulated_interest(loan, as_of_date=None):
    """Calculate total accumulated interest for a loan - returns both daily and monthly calculations"""
    try:
        if as_of_date is None:
            as_of_date = date.today()
        
        # Calculate days since loan creation
        days_since_creation = (as_of_date - loan.created_at.date()).days
        
        # Get verified interest payments (common for both calculations)
        verified_interest_payments = get_payment_query().with_entities(db.func.sum(Payment.interest_amount)).filter_by(
            loan_id=loan.id, 
            status='verified'
        ).scalar() or 0
        verified_interest_payments = Decimal(str(verified_interest_payments))
        
        # Calculate daily accumulated interest
        if loan.loan_type == 'interest_only':
            # For interest-only loans, calculate interest on original principal
            daily_total_interest = calculate_interest_for_period(
                loan.principal_amount, 
                loan.interest_rate, 
                loan.created_at.date(), 
                as_of_date
            )
        else:
            # For regular loans, use daily calculation
            daily_total_interest = calculate_interest_for_period(
                loan.principal_amount, 
                loan.interest_rate, 
                loan.created_at.date(), 
                as_of_date
            )
        
        daily_accumulated_interest = daily_total_interest - verified_interest_payments
        
        # Calculate monthly accumulated interest
        if days_since_creation < 30:
            # Before 30 days: return 0 for monthly calculation
            monthly_accumulated_interest = Decimal('0')
        else:
            # After 30 days: calculate monthly interest
            months_passed = days_since_creation // 30
            monthly_interest = calculate_monthly_interest(loan.principal_amount, loan.interest_rate)
            monthly_total_interest = monthly_interest * months_passed
            monthly_accumulated_interest = monthly_total_interest - verified_interest_payments
        
        # Return both calculations
        return {
            'daily': daily_accumulated_interest,
            'monthly': monthly_accumulated_interest,
            'days_since_creation': days_since_creation,
            'months_passed': days_since_creation // 30 if days_since_creation >= 30 else 0
        }
        
    except Exception as e:
        print(f"Error calculating accumulated interest: {e}")
        return {
            'daily': Decimal('0'),
            'monthly': Decimal('0'),
            'days_since_creation': 0,
            'months_passed': 0
        }

def verify_payment(payment_id):
    """Verify a payment and update loan balance"""
    try:
        payment = get_payment_query().filter_by(id=payment_id).first()
        if not payment:
            return  # Payment not found
        loan = payment.loan
        
        if payment.status == 'verified':
            return  # Already verified
        
        # Update payment status
        payment.status = 'verified'
        
        # Update loan balance (only for regular loans)
        if loan.loan_type != 'interest_only':
            loan.remaining_principal -= payment.principal_amount
            if loan.remaining_principal < 0:
                loan.remaining_principal = Decimal('0')
        
        commit_current_instance()
        
    except Exception as e:
        # Rollback is handled by the instance-specific session
        raise e

def generate_loan_calculation_excel(loan):
    """Generate Excel report showing daily loan calculations for 6 months"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    
    # Sheet 1: With Payments - Database (Static from DB)
    ws1 = wb.active
    ws1.title = "DB Payments (Actual)"
    
    # Sheet 2: With Payments - Manual Entry (Formulas)
    ws2 = wb.create_sheet("Manual Payments")
    
    # Sheet 3: Without Payments (Projection)
    ws3 = wb.create_sheet("Without Payments")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all payments for this loan (sorted by date)
    payments = get_payment_query().filter_by(loan_id=loan.id, status='verified').order_by(Payment.payment_date).all()
    
    # Calculate for 1 year (365 days)
    num_days = 365
    start_date = loan.created_at.date()
    
    def setup_sheet(ws, include_payments=True):
        """Setup headers and styling for a sheet"""
        headers = [
            "Day", "Date", "Opening Principal", "Daily Interest", 
            "Accumulated Interest", "Payment Amount", "Interest Paid", 
            "Principal Paid", "Closing Principal", "Remaining Interest"
        ]
        
        if not include_payments:
            headers = [h for h in headers if h not in ["Payment Amount", "Interest Paid", "Principal Paid"]]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        return headers
    
    # Setup all three sheets
    headers_db = setup_sheet(ws1, include_payments=True)
    headers_manual = setup_sheet(ws2, include_payments=True)
    headers_without = setup_sheet(ws3, include_payments=False)
    
    # Get all payments for sheets with database data
    # Handle interest rate stored as decimal (0.12) or percentage (12)
    interest_rate_value = float(loan.interest_rate)
    if interest_rate_value < 1:
        # Stored as decimal (0.12 for 12%), don't divide by 100
        daily_rate = interest_rate_value / DAYS_PER_YEAR
    else:
        # Stored as percentage (12 for 12%), divide by 100
        daily_rate = interest_rate_value / DAYS_PER_YEAR / 100
    
    # Pre-populate payment amounts from database
    payment_dict = {}
    for payment in payments:
        payment_date_key = payment.payment_date.date()
        if payment_date_key not in payment_dict:
            payment_dict[payment_date_key] = {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')}
        payment_dict[payment_date_key]['amount'] += payment.amount
        payment_dict[payment_date_key]['interest'] += payment.interest_amount
        payment_dict[payment_date_key]['principal'] += payment.principal_amount
    
    # ===== SHEET 1: Database Payments (Static from DB) =====
    opening_principal = loan.principal_amount
    accumulated_interest = Decimal('0')
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2
        
        # Daily interest on opening principal
        daily_interest = opening_principal * Decimal(str(daily_rate))
        accumulated_interest += daily_interest
        
        # Get payments from database for this date
        payment_data = payment_dict.get(current_date, {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')})
        payment_amount = payment_data['amount']
        interest_paid = payment_data['interest']
        principal_paid = payment_data['principal']
        
        # Update accumulated interest and principal
        accumulated_interest -= interest_paid
        closing_principal = opening_principal - principal_paid
        
        # Write data to sheet (static values from database)
        ws1.cell(row=row, column=1, value=day + 1)
        ws1.cell(row=row, column=2, value=current_date.strftime('%Y-%m-%d'))
        ws1.cell(row=row, column=3, value=float(opening_principal)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=4, value=float(daily_interest)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=5, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=6, value=float(payment_amount)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=7, value=float(interest_paid)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=8, value=float(principal_paid)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=9, value=float(closing_principal)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=10, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        
        # Apply borders
        for col in range(1, 11):
            ws1.cell(row=row, column=col).border = border
        
        # Highlight payment rows
        if payment_amount > 0:
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Update for next day
        opening_principal = closing_principal
    
    # ===== SHEET 2: Manual Payments (Formulas for manual entry) =====
    # This sheet has formulas so you can manually enter payments and see calculations update
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2  # Starting at row 2 for data (headers will be inserted later, shifting to row 7)
        final_row = row + 5  # Account for 5 header rows that will be inserted
        
        # Get pre-populated payment data
        payment_data = payment_dict.get(current_date, {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')})
        
        # Write static values
        ws2.cell(row=row, column=1, value=day + 1)  # Day
        ws2.cell(row=row, column=2, value=current_date.strftime('%Y-%m-%d'))  # Date
        
        # Opening Principal (formula: previous day's closing or initial value)
        if day == 0:
            # First day: use initial principal as value
            ws2.cell(row=row, column=3, value=float(loan.principal_amount)).number_format = '₹#,##0.00'
        else:
            # Subsequent days: reference previous day's closing with error handling
            ws2.cell(row=row, column=3, value=f'=IFERROR(I{final_row-1},{float(loan.principal_amount)})').number_format = '₹#,##0.00'
        
        # Daily Interest (formula: Opening Principal * daily rate)
        ws2.cell(row=row, column=4, value=f'=IFERROR(C{final_row}*{daily_rate},0)').number_format = '₹#,##0.00'
        
        # Payment Amount (editable - pre-populate from database) - MOVED BEFORE accumulated interest
        ws2.cell(row=row, column=6, value=float(payment_data['amount'])).number_format = '₹#,##0.00'
        
        # Accumulated Interest BEFORE payment (formula: previous remaining + today's interest)
        if day == 0:
            # First day: just today's interest
            ws2.cell(row=row, column=5, value=f'=IFERROR(D{final_row},0)').number_format = '₹#,##0.00'
        else:
            # Subsequent days: previous day's REMAINING interest (J) + today's new interest
            ws2.cell(row=row, column=5, value=f'=IFERROR(J{final_row-1}+D{final_row},0)').number_format = '₹#,##0.00'
        
        # Interest Paid (formula: MIN(payment amount, accumulated interest before payment))
        ws2.cell(row=row, column=7, value=f'=IFERROR(MIN(F{final_row},E{final_row}),0)').number_format = '₹#,##0.00'
        
        # Principal Paid (formula: payment amount - interest paid)
        ws2.cell(row=row, column=8, value=f'=IFERROR(F{final_row}-G{final_row},0)').number_format = '₹#,##0.00'
        
        # Closing Principal (formula: opening principal - principal paid)
        ws2.cell(row=row, column=9, value=f'=IFERROR(C{final_row}-H{final_row},0)').number_format = '₹#,##0.00'
        
        # Remaining Interest (formula: accumulated interest AFTER payment = accumulated - interest paid)
        ws2.cell(row=row, column=10, value=f'=IFERROR(E{final_row}-G{final_row},0)').number_format = '₹#,##0.00'
        
        # Apply borders
        for col in range(1, 11):
            ws2.cell(row=row, column=col).border = border
        
        # Highlight payment rows (from database - can be cleared for manual entry)
        if payment_data['amount'] > 0:
            for col in range(1, 11):
                ws2.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # ===== SHEET 3: Without Payments (Projection) =====
    # Static projection showing interest accumulation without any payments
    opening_principal = loan.principal_amount
    accumulated_interest = Decimal('0')
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2
        
        # Daily interest on opening principal (no reduction since no payments)
        daily_interest = opening_principal * Decimal(str(daily_rate))
        accumulated_interest += daily_interest
        
        # Write data to sheet
        col = 1
        ws3.cell(row=row, column=col, value=day + 1)
        col += 1
        ws3.cell(row=row, column=col, value=current_date.strftime('%Y-%m-%d'))
        col += 1
        ws3.cell(row=row, column=col, value=float(opening_principal)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(daily_interest)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(opening_principal)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        
        # Apply borders
        for c in range(1, len(headers_without) + 1):
            ws3.cell(row=row, column=c).border = border
    
    # Add summary information at the top of all sheets
    for ws in [ws1, ws2, ws3]:
        ws.insert_rows(1, 5)
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Loan Calculation Report: {loan.loan_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Handle interest rate display (stored as decimal or percentage)
        interest_rate_display = float(loan.interest_rate)
        if interest_rate_display < 1:
            interest_rate_display = interest_rate_display * 100
        
        ws['A2'] = f"Customer: {loan.customer.username}"
        ws['A3'] = f"Principal Amount: ₹{float(loan.principal_amount):,.2f}"
        ws['A4'] = f"Interest Rate: {interest_rate_display:.2f}% per annum"
        ws['A5'] = f"Loan Created: {loan.created_at.strftime('%Y-%m-%d')}"
        
        # Move headers down
        for row in ws.iter_rows(min_row=6, max_row=6):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Re-apply column widths after inserting rows (to ensure they're preserved)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Routes
@app.before_request
def before_request():
    """Configure app for current instance before each request"""
    instance = get_current_instance()
    g.current_instance = instance
    
    # Just store the instance in g - don't modify the global db object
    # The helper functions will handle instance-specific database access
    if instance not in VALID_INSTANCES:
        flash('Invalid instance', 'error')
        return redirect('/')

@app.context_processor
def inject_current_user():
    """Make current_user, cashback_balance, and current_instance available in all templates"""
    from flask_login import current_user
    instance = getattr(g, 'current_instance', 'prod')
    cashback_balance = Decimal('0')
    
    try:
        if current_user.is_authenticated:
            cashback_balance = get_user_cashback_balance(current_user.id, instance)
    except Exception:
        pass
    
    return dict(
        current_user=current_user,
        cashback_balance=cashback_balance,
        current_instance=instance
    )

@app.template_filter('from_json')
def from_json_filter(value):
    """Parse JSON string to Python object"""
    if not value:
        return {}
    try:
        return json.loads(value)
    except:
        return {}

@app.route('/')
def index():
    """Default route - redirect to production instance"""
    return redirect(url_for('login', instance_name='prod'))

@app.route('/instances')
def instances():
    """Instance selector page - show all instances"""
    instances_info = {}
    for instance in VALID_INSTANCES:
        instances_info[instance] = {
            'name': instance,
            'database_exists': Path(f"instances/{instance}/database/lending_app_{instance}.db").exists(),
            'database_size': 0
        }
        
        db_path = Path(f"instances/{instance}/database/lending_app_{instance}.db")
        if db_path.exists():
            instances_info[instance]['database_size'] = db_path.stat().st_size
    
    return render_template('instance_selector.html', instances_info=instances_info)

@app.route('/login_redirect')
def login_redirect():
    """Redirect to login for current instance"""
    next_url = request.args.get('next', '')
    if next_url:
        # Extract instance from next URL
        if next_url.startswith('/'):
            next_url = next_url[1:]  # Remove leading slash
        path_parts = next_url.split('/')
        if path_parts and path_parts[0] in VALID_INSTANCES:
            instance = path_parts[0]
        else:
            instance = DEFAULT_INSTANCE
    else:
        instance = get_current_instance()
    return redirect(url_for('login', instance_name=instance))

@app.route('/<instance_name>/')
def instance_index(instance_name):
    """Redirect to instance login"""
    if instance_name in VALID_INSTANCES:
        return redirect(f'/{instance_name}/login')
    else:
        return redirect('/')


# ============================================================================
# DECORATORS
# ============================================================================

# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/<instance_name>/login', methods=['GET', 'POST'])
def login(instance_name):
    """User login for specific instance"""
    import logging
    logger = logging.getLogger(__name__)
    
    # DEBUG: Log login attempt
    print(f"[DEBUG LOGIN] Login attempt - instance_name: {instance_name}")
    print(f"[DEBUG LOGIN] VALID_INSTANCES: {VALID_INSTANCES}")
    print(f"[DEBUG LOGIN] Request method: {request.method}")
    
    if instance_name not in VALID_INSTANCES:
        print(f"[DEBUG LOGIN] Invalid instance: {instance_name}, redirecting to /")
        return redirect('/')
    
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        
        print(f"[DEBUG LOGIN] POST request received")
        print(f"[DEBUG LOGIN] Username: {username}")
        print(f"[DEBUG LOGIN] Password provided: {'Yes' if password else 'No'}")
        
        try:
            # Get database URI for debugging
            db_uri = get_database_uri(instance_name)
            print(f"[DEBUG LOGIN] Database URI: {db_uri}")
            
            # Get user query
            user_query = get_user_query()
            print(f"[DEBUG LOGIN] User query object: {user_query}")
            
            user = user_query.filter_by(username=username).first()
            print(f"[DEBUG LOGIN] User found: {user is not None}")
            
            if user:
                print(f"[DEBUG LOGIN] User ID: {user.id}")
                print(f"[DEBUG LOGIN] User username: {user.username}")
                print(f"[DEBUG LOGIN] User password_hash exists: {hasattr(user, 'password_hash') and user.password_hash is not None}")
                print(f"[DEBUG LOGIN] User is_admin: {user.is_admin}")
                print(f"[DEBUG LOGIN] User is_moderator: {user.is_moderator}")
            
            if user and check_password_hash(user.password_hash, password):
                print(f"[DEBUG LOGIN] Password check passed")
                try:
                    # Ensure we're in the correct app context
                    from flask import current_app
                    if not hasattr(current_app, 'login_manager'):
                        print("[DEBUG LOGIN] ERROR: login_manager not found in current_app")
                        flash('Login system not properly initialized. Please try again.', 'error')
                        return render_template('login.html', instance_name=instance_name)
                    
                    print(f"[DEBUG LOGIN] Calling login_user for user: {user.username}")
                    login_user(user)
                    print(f"[DEBUG LOGIN] login_user successful")
                    
                    # Log successful login
                    try:
                        logging_mgr = get_logging_manager(instance_name)
                        metrics_mgr = get_metrics_manager(instance_name)
                        logging_mgr.log_login(user.username, success=True)
                        metrics_mgr.record_login(user.username, success=True)
                    except Exception as log_error:
                        print(f"[DEBUG LOGIN] Error logging login: {log_error}")
                except Exception as e:
                    print(f"[DEBUG LOGIN] ERROR during login_user: {e}")
                    import traceback
                    traceback.print_exc()
                    flash('Login failed. Please try again.', 'error')
                    return render_template('login.html', instance_name=instance_name)
                
                next_page = request.args.get('next')
                print(f"[DEBUG LOGIN] Next page: {next_page}")
                
                if user.is_admin:
                    redirect_url = next_page if next_page else url_for('admin_dashboard', instance_name=instance_name)
                    print(f"[DEBUG LOGIN] Redirecting admin to: {redirect_url}")
                    return redirect(redirect_url)
                elif user.is_moderator or len(user.assigned_loans.filter_by(is_active=True).all()) > 0 or len(user.assigned_trackers.filter_by(is_active=True).all()) > 0:
                    # User is a moderator - redirect to customer dashboard (they can access moderator view from there)
                    redirect_url = next_page if next_page else url_for('customer_dashboard', instance_name=instance_name)
                    print(f"[DEBUG LOGIN] Redirecting moderator to customer dashboard: {redirect_url}")
                    return redirect(redirect_url)
                else:
                    redirect_url = next_page if next_page else url_for('customer_dashboard', instance_name=instance_name)
                    print(f"[DEBUG LOGIN] Redirecting customer to: {redirect_url}")
                    return redirect(redirect_url)
            else:
                print(f"[DEBUG LOGIN] Password check failed or user not found")
                if not user:
                    print(f"[DEBUG LOGIN] User '{username}' not found in database")
                else:
                    print(f"[DEBUG LOGIN] Password mismatch for user '{username}'")
                
                # Log failed login attempt
                try:
                    logging_mgr = get_logging_manager(instance_name)
                    metrics_mgr = get_metrics_manager(instance_name)
                    reason = 'User not found' if not user else 'Invalid password'
                    logging_mgr.log_login(username, success=False, reason=reason)
                    metrics_mgr.record_login(username, success=False)
                except Exception as log_error:
                    print(f"[DEBUG LOGIN] Error logging failed login: {log_error}")
                
                flash('Invalid username or password')
        except Exception as e:
            print(f"[DEBUG LOGIN] EXCEPTION in login route: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Login error: {str(e)}', 'error')
    
    print(f"[DEBUG LOGIN] Rendering login template for instance: {instance_name}")
    return render_template('login.html', instance_name=instance_name)

@app.route('/<instance_name>/logout')
@login_required
def logout(instance_name):
    """User logout"""
    username = current_user.username if hasattr(current_user, 'username') else 'unknown'
    
    # Log logout
    try:
        logging_mgr = get_logging_manager(instance_name)
        metrics_mgr = get_metrics_manager(instance_name)
        logging_mgr.log_logout(username)
        metrics_mgr.record_logout(username)
    except Exception as log_error:
        print(f"[DEBUG LOGOUT] Error logging logout: {log_error}")
    
    logout_user()
    return redirect(url_for('login', instance_name=instance_name))


# ============================================================================
# Language Switching Routes (i18n)
# ============================================================================

@app.route('/<instance_name>/set-language/<language_code>')
def set_language(instance_name, language_code):
    """
    Set user's language preference
    
    For logged-in users: Saves to database
    For guests: Saves to session
    """
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    supported_languages = get_supported_languages()
    
    if language_code not in supported_languages:
        flash(gettext('Invalid language selection'), 'error')
        return redirect(request.referrer or url_for('login', instance_name=instance_name))
    
    # Save to session for immediate effect
    session['language'] = language_code
    
    # If user is logged in, save to database
    if current_user.is_authenticated:
        try:
            current_user.language_preference = language_code
            db.session.commit()
            flash(gettext('Language preference saved'), 'success')
        except Exception as e:
            db.session.rollback()
            flash(gettext('Error saving language preference'), 'error')
    
    # Redirect back to referring page or login
    return redirect(request.referrer or url_for('login', instance_name=instance_name))


@app.route('/<instance_name>/register', methods=['GET', 'POST'])
def register(instance_name):
    """User registration for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form.get('email', '')  # Email is optional
        password = request.form['password']
        
        if get_user_query().filter_by(username=username).first():
            flash('Username already exists')
            return render_template('register.html', instance_name=instance_name)
        
        user = User(
            username=username,
            email=email if email else None,
            password_hash=generate_password_hash(password),
            is_admin=False
        )
        
        add_to_current_instance(user)
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login', instance_name=instance_name))
    
    return render_template('register.html', instance_name=instance_name)

# Admin routes
@app.route('/<instance_name>/admin')
@login_required
def admin_dashboard(instance_name):
    """Admin dashboard for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    # Get statistics (exclude closed loans)
    total_loans = get_loan_query().filter_by(status='active', is_active=True).count()
    total_principal = sum(loan.principal_amount for loan in get_loan_query().filter_by(status='active', is_active=True).all())
    total_interest_earned = sum(payment.interest_amount for payment in get_payment_query().filter_by(status='verified').all())
    total_users = get_user_query().count()
    tracker_summary_errors = []

    trackers = get_daily_tracker_query().all()
    active_trackers = [tracker for tracker in trackers if tracker.is_active]
    total_active_trackers = len(active_trackers)

    total_tracker_investment = Decimal('0')
    total_tracker_returns = Decimal('0')

    for tracker in active_trackers:
        investment_value = tracker.investment or Decimal('0')
        if not isinstance(investment_value, Decimal):
            investment_value = Decimal(str(investment_value))
        total_tracker_investment += investment_value

        if tracker.filename:
            try:
                summary = get_tracker_summary(instance_name, tracker.filename)
                payments_value = summary.get('total_payments') or 0
                total_tracker_returns += Decimal(str(payments_value))
            except Exception as e:
                tracker_summary_errors.append((tracker.id, str(e)))

    total_tracker_investment_float = float(total_tracker_investment) if total_tracker_investment else 0.0
    total_tracker_returns_float = float(total_tracker_returns) if total_tracker_returns else 0.0
    
    # Calculate loan-wise cashback (total cashback from all loans)
    loan_cashback_total = Decimal('0')
    from app_loans import get_loan_cashback_total
    from app_trackers import get_tracker_cashback_total
    all_loans = get_loan_query().all()
    for loan in all_loans:
        loan_cashback = get_loan_cashback_total(loan.id, instance_name)
        loan_cashback_total += loan_cashback
    loan_cashback_total_float = float(loan_cashback_total) if loan_cashback_total else 0.0
    
    # Calculate tracker-wise cashback (total cashback from all trackers)
    tracker_cashback_total = Decimal('0')
    for tracker in active_trackers:
        tracker_cashback = get_tracker_cashback_total(tracker.id, instance_name)
        tracker_cashback_total += tracker_cashback
    tracker_cashback_total_float = float(tracker_cashback_total) if tracker_cashback_total else 0.0
    
    # Get pending payments count
    pending_payments_count = get_payment_query().filter_by(status='pending').count()
    
    # Get pending tracker entries count
    from app_trackers import TrackerEntry, get_tracker_entry_query
    pending_tracker_entries_count = get_tracker_entry_query().filter_by(status='pending').count()
    
    return render_template('admin/dashboard.html', 
                         total_loans=total_loans,
                         total_principal=total_principal,
                         total_interest_earned=total_interest_earned,
                         total_users=total_users,
                         total_active_trackers=total_active_trackers,
                         total_tracker_investment=total_tracker_investment_float,
                         total_tracker_returns=total_tracker_returns_float,
                         loan_cashback_total=loan_cashback_total_float,
                         tracker_cashback_total=tracker_cashback_total_float,
                         tracker_summary_errors=tracker_summary_errors,
                         pending_payments_count=pending_payments_count,
                         pending_tracker_entries_count=pending_tracker_entries_count,
                         instance_name=instance_name)

# Admin Activity Logs route
@app.route('/<instance_name>/admin/activity-logs')
@login_required
def admin_activity_logs(instance_name):
    """Admin activity logs page"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    from datetime import datetime, timedelta
    from lms_logging import get_logging_manager, ActivityLog
    from sqlalchemy.orm import sessionmaker
    
    # Get date filter
    date_filter = request.args.get('date', 'today')
    if date_filter == 'today':
        start_date = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = datetime.utcnow()
    elif date_filter == 'week':
        start_date = datetime.utcnow() - timedelta(days=7)
        end_date = datetime.utcnow()
    elif date_filter == 'month':
        start_date = datetime.utcnow() - timedelta(days=30)
        end_date = datetime.utcnow()
    else:
        start_date = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = datetime.utcnow()
    
    # Get filter parameters
    action_filter = request.args.get('action', '')
    username_filter = request.args.get('username', '')
    
    logging_mgr = get_logging_manager(instance_name)
    engine = db_manager.get_engine_for_instance(instance_name)
    Session = sessionmaker(bind=engine)
    session = Session()
    
    try:
        # Query activity logs
        query = session.query(ActivityLog).filter(
            ActivityLog.created_at >= start_date,
            ActivityLog.created_at <= end_date
        )
        
        if action_filter:
            query = query.filter(ActivityLog.action == action_filter)
        if username_filter:
            query = query.filter(ActivityLog.username == username_filter)
        
        logs = query.order_by(ActivityLog.created_at.desc()).limit(500).all()
        
        # Get unique actions and usernames for filters
        all_actions = session.query(ActivityLog.action).distinct().all()
        all_usernames = session.query(ActivityLog.username).distinct().all()
        
        # Get today's summary
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        today_logins = session.query(ActivityLog).filter_by(action='login_success').filter(
            ActivityLog.created_at >= today_start
        ).count()
        today_logouts = session.query(ActivityLog).filter_by(action='logout').filter(
            ActivityLog.created_at >= today_start
        ).count()
        today_payments = session.query(ActivityLog).filter(
            ActivityLog.action.like('%payment%')
        ).filter(
            ActivityLog.created_at >= today_start
        ).count()
        today_cashback = session.query(ActivityLog).filter(
            ActivityLog.action.like('%cashback%')
        ).filter(
            ActivityLog.created_at >= today_start
        ).count()
        today_admin_actions = session.query(ActivityLog).filter(
            ActivityLog.action.like('admin_%')
        ).filter(
            ActivityLog.created_at >= today_start
        ).count()
        
    finally:
        session.close()
    
    # Get timezone info
    timezone_str = logging_mgr.get_config('system_timezone', 'Asia/Kolkata')
    
    return render_template('admin/activity_logs.html',
                         logs=logs,
                         date_filter=date_filter,
                         action_filter=action_filter,
                         username_filter=username_filter,
                         all_actions=[a[0] for a in all_actions],
                         all_usernames=[u[0] for u in all_usernames],
                         today_logins=today_logins,
                         today_logouts=today_logouts,
                         today_payments=today_payments,
                         today_cashback=today_cashback,
                         today_admin_actions=today_admin_actions,
                         logging_mgr=logging_mgr,
                         timezone_str=timezone_str,
                         instance_name=instance_name)

# Admin Metrics Dashboard route
@app.route('/<instance_name>/admin/metrics')
@login_required
def admin_metrics(instance_name):
    """Admin metrics dashboard"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    from datetime import date, timedelta, datetime
    from lms_metrics import get_metrics_manager
    from lms_logging import get_logging_manager
    from decimal import Decimal
    
    # Get period filter
    period = request.args.get('period', 'today')
    
    metrics_mgr = get_metrics_manager(instance_name)
    logging_mgr = get_logging_manager(instance_name)
    
    # Get threshold for overdue interest
    threshold_days = int(logging_mgr.get_config('interest_payment_threshold_days', '30'))
    
    # Calculate date ranges
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Get payment metrics from main database
    from sqlalchemy import func
    engine = db_manager.get_engine_for_instance(instance_name)
    Session = sessionmaker(bind=engine)
    session = Session()
    
    try:
        # Get pending payments (all time)
        pending_payments_query = get_payment_query().filter_by(status='pending')
        pending_payments_count = pending_payments_query.count()
        pending_payments_amount = pending_payments_query.with_entities(
            func.sum(Payment.amount)
        ).scalar() or Decimal('0')
        
        # Calculate today's payment metrics
        today_start = datetime.combine(today, datetime.min.time())
        today_payments = get_payment_query().filter(
            Payment.payment_date >= today_start
        ).all()
        today_pending = [p for p in today_payments if p.status == 'pending']
        today_verified = [p for p in today_payments if p.status == 'verified']
        today_payment_metrics = {
            'pending': {
                'count': len(today_pending),
                'total': float(sum(p.amount for p in today_pending))
            },
            'verified': {
                'count': len(today_verified),
                'total': float(sum(p.amount for p in today_verified))
            }
        }
        
        # Calculate week's payment metrics
        week_start = datetime.combine(week_start, datetime.min.time())
        week_payments = get_payment_query().filter(
            Payment.payment_date >= week_start
        ).all()
        week_pending = [p for p in week_payments if p.status == 'pending']
        week_verified = [p for p in week_payments if p.status == 'verified']
        week_payment_metrics = {
            'pending': {
                'count': len(week_pending),
                'total': float(sum(p.amount for p in week_pending))
            },
            'verified': {
                'count': len(week_verified),
                'total': float(sum(p.amount for p in week_verified))
            }
        }
        
        # Calculate month's payment metrics
        month_start = datetime.combine(month_start, datetime.min.time())
        month_payments = get_payment_query().filter(
            Payment.payment_date >= month_start
        ).all()
        month_pending = [p for p in month_payments if p.status == 'pending']
        month_verified = [p for p in month_payments if p.status == 'verified']
        month_payment_metrics = {
            'pending': {
                'count': len(month_pending),
                'total': float(sum(p.amount for p in month_pending))
            },
            'verified': {
                'count': len(month_verified),
                'total': float(sum(p.amount for p in month_verified))
            }
        }
        
        # Get loans with overdue interest
        all_loans = get_loan_query().filter_by(is_active=True, status='active').all()
        overdue_loans = []
        
        for loan in all_loans:
            # Calculate days since last payment
            last_payment = get_payment_query().filter_by(
                loan_id=loan.id,
                status='verified'
            ).order_by(Payment.payment_date.desc()).first()
            
            if last_payment:
                days_since_payment = (date.today() - last_payment.payment_date.date()).days
            else:
                # No verified payments, use loan creation date
                days_since_payment = (date.today() - loan.created_at.date()).days
            
            if days_since_payment > threshold_days:
                # Calculate accumulated interest based on loan frequency
                interest_data = calculate_accumulated_interest(loan, date.today())
                
                # Use monthly interest for monthly loans, daily for daily loans
                if loan.payment_frequency == 'monthly':
                    accumulated_interest = interest_data.get('monthly', Decimal('0'))
                else:  # daily
                    accumulated_interest = interest_data.get('daily', Decimal('0'))
                
                overdue_loans.append({
                    'loan': loan,
                    'days_overdue': days_since_payment,
                    'accumulated_interest': accumulated_interest,
                    'loan_type': loan.payment_frequency  # 'monthly' or 'daily'
                })
        
        # Sort by days overdue (most overdue first)
        overdue_loans.sort(key=lambda x: x['days_overdue'], reverse=True)
        
        # Get trackers with pending > 0 for more than configured threshold days
        # tracker_manager is already imported at the top of the file
        all_trackers = get_daily_tracker_query().filter_by(is_active=True).all()
        overdue_trackers = []
        tracker_pending_threshold = int(logging_mgr.get_config('tracker_pending_threshold_days', '5'))
        
        for tracker in all_trackers:
            try:
                tracker_data = get_tracker_data(instance_name, tracker.filename)
                summary = get_tracker_summary(instance_name, tracker.filename)
                pending = float(summary.get('pending', 0))
                
                # Only consider trackers with positive pending (outstanding payments)
                if pending > 0:
                    # Find last paid date
                    last_paid_date = None
                    for row in reversed(tracker_data['data']):
                        daily_payment = row.get('daily_payments', 0)
                        if daily_payment and float(daily_payment or 0) > 0:
                            row_date = row.get('date')
                            if row_date:
                                try:
                                    # Date can be date object, datetime object, or string
                                    if isinstance(row_date, date):
                                        last_paid_date = row_date
                                    elif isinstance(row_date, datetime):
                                        last_paid_date = row_date.date()
                                    elif isinstance(row_date, str):
                                        last_paid_date = datetime.strptime(row_date, '%Y-%m-%d').date()
                                    break
                                except Exception as e:
                                    # Skip invalid dates
                                    pass
                    
                    # Calculate days since last payment
                    if last_paid_date:
                        days_since_payment = (date.today() - last_paid_date).days
                    else:
                        # No payment found, use tracker creation date
                        days_since_payment = (date.today() - tracker.created_at.date()).days
                    
                    # If pending > 0 and more than 5 days since last payment
                    if days_since_payment > tracker_pending_threshold:
                        overdue_trackers.append({
                            'tracker': tracker,
                            'pending': pending,
                            'days_since_payment': days_since_payment,
                            'last_paid_date': last_paid_date,
                            'total_payments': summary.get('total_payments', 0),
                            'expected_total': summary.get('expected_total', 0)
                        })
            except Exception as e:
                # Skip trackers with errors
                print(f"Error processing tracker {tracker.id}: {e}")
                continue
        
        # Sort by days since payment (most overdue first)
        overdue_trackers.sort(key=lambda x: x['days_since_payment'], reverse=True)
        
    finally:
        session.close()
    
    return render_template('admin/metrics.html',
                         period=period,
                         threshold_days=threshold_days,
                         today_pending_count=pending_payments_count,
                         today_pending_amount=float(pending_payments_amount),
                         today_payment_metrics=today_payment_metrics,
                         week_payment_metrics=week_payment_metrics,
                         month_payment_metrics=month_payment_metrics,
                         overdue_loans=overdue_loans,
                         overdue_trackers=overdue_trackers,
                         tracker_pending_threshold=tracker_pending_threshold,
                         instance_name=instance_name)

# Admin System Configuration route
@app.route('/<instance_name>/admin/config', methods=['GET', 'POST'])
@login_required
def admin_config(instance_name):
    """Admin system configuration page"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    from lms_logging import get_logging_manager
    
    logging_mgr = get_logging_manager(instance_name)
    
    if request.method == 'POST':
        threshold_days = request.form.get('interest_payment_threshold_days', '30')
        timezone = request.form.get('system_timezone', 'Asia/Kolkata')
        tracker_threshold = request.form.get('tracker_pending_threshold_days', '5')
        
        try:
            logging_mgr.set_config(
                'interest_payment_threshold_days',
                threshold_days,
                description='Days threshold for highlighting overdue interest payments',
                updated_by=current_user.id
            )
            logging_mgr.set_config(
                'system_timezone',
                timezone,
                description='System timezone for displaying timestamps',
                updated_by=current_user.id
            )
            logging_mgr.set_config(
                'tracker_pending_threshold_days',
                tracker_threshold,
                description='Days threshold for highlighting trackers with pending payments',
                updated_by=current_user.id
            )
            flash('Configuration updated successfully', 'success')
        except Exception as e:
            flash(f'Error updating configuration: {str(e)}', 'error')
        
        return redirect(url_for('admin_config', instance_name=instance_name))
    
    # Get current config
    threshold = logging_mgr.get_config('interest_payment_threshold_days', '30')
    timezone = logging_mgr.get_config('system_timezone', 'Asia/Kolkata')
    tracker_threshold = logging_mgr.get_config('tracker_pending_threshold_days', '5')
    
    return render_template('admin/config.html',
                         threshold=threshold,
                         timezone=timezone,
                         tracker_threshold=tracker_threshold,
                         instance_name=instance_name)

# Admin Cashback Management Routes
# Loan routes moved to app_loans.py
# Tracker routes moved to app_trackers.py

# Admin Users route
@app.route('/<instance_name>/admin/users')
@login_required
def admin_users(instance_name):
    """Admin users page for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    users = get_user_query().all()
    total_loans = sum(len(user.loans) for user in users)
    admin_count = sum(1 for user in users if user.is_admin)
    moderator_count = sum(1 for user in users if user.is_moderator and not user.is_admin)
    customer_count = len(users) - admin_count - moderator_count
    
    # Calculate cashback balance for each user
    users_with_balance = []
    for user in users:
        balance = get_user_cashback_balance(user.id, instance_name)
        users_with_balance.append({
            'user': user,
            'cashback_balance': balance
        })
    
    return render_template('admin/users.html', 
                         users_with_balance=users_with_balance,
                         users=users,  # Keep for backward compatibility
                         total_loans=total_loans,
                         admin_count=admin_count,
                         moderator_count=moderator_count,
                         customer_count=customer_count,
                         instance_name=instance_name)

# Admin Edit User Email route
@app.route('/<instance_name>/admin/user/<int:user_id>/edit-email', methods=['POST'])
@login_required
def admin_edit_user_email(instance_name, user_id):
    """Admin edit user email"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    user = get_user_query().get(user_id)
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    new_email = request.form.get('email', '').strip()
    
    # Allow empty email (set to None)
    if not new_email:
        user.email = None
        flash(f'Email removed for user {user.username}', 'success')
    else:
        # Validate email format (basic validation)
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, new_email):
            flash('Invalid email format', 'error')
            return redirect(url_for('admin_users', instance_name=instance_name))
        
        # Check if email already exists for another user
        existing_user = get_user_query().filter_by(email=new_email).first()
        if existing_user and existing_user.id != user_id:
            flash(f'Email already in use by user: {existing_user.username}', 'error')
            return redirect(url_for('admin_users', instance_name=instance_name))
        
        user.email = new_email
        flash(f'Email updated for user {user.username}', 'success')
    
    commit_current_instance()
    
    return redirect(url_for('admin_users', instance_name=instance_name))

# Admin Create User route
@app.route('/<instance_name>/admin/create-user', methods=['GET', 'POST'])
@login_required
def admin_create_user(instance_name):
    """Admin create user page for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form.get('email', '')  # Email is optional
        password = request.form['password']
        is_admin = 'is_admin' in request.form
        
        if get_user_query().filter_by(username=username).first():
            flash('Username already exists')
            return render_template('admin/create_user.html', instance_name=instance_name)
        
        user = User(
            username=username,
            email=email if email else None,
            password_hash=generate_password_hash(password),
            is_admin=is_admin
        )
        add_to_current_instance(user)
        
        flash(f'User {username} created successfully')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    return render_template('admin/create_user.html', instance_name=instance_name)

# Admin Payments route (kept here as it's not loan-specific)
# Loan routes moved to app_loans.py - removed duplicates

# Admin Create Backup route
# Customer routes
# Admin Payments route
@app.route('/<instance_name>/admin/payments')
@login_required
def admin_payments(instance_name):
    """Admin payments page for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    # Get filter parameters
    customer_filter = request.args.get('customer', '')
    loan_filter = request.args.get('loan', '')
    status_filter = request.args.get('status', '')
    payment_method = request.args.get('payment_method', '')
    
    # Build query
    # Get the current instance session for complex queries
    instance = get_current_instance_from_g()
    session = db_manager.get_session_for_instance(instance)
    query = session.query(Payment, Loan, User).join(Loan, Payment.loan_id == Loan.id).join(User, Loan.customer_id == User.id)
    
    if customer_filter:
        query = query.filter(User.username.contains(customer_filter))
    if loan_filter:
        query = query.filter(Loan.loan_name.contains(loan_filter))
    if status_filter:
        query = query.filter(Payment.status == status_filter)
    if payment_method:
        query = query.filter(Payment.payment_method == payment_method)
    
    # Get sorting parameters
    sort_by = request.args.get('sort', 'payment_date')
    sort_order = request.args.get('order', 'desc')
    
    if sort_by == 'customer':
        if sort_order == 'asc':
            query = query.order_by(User.username.asc())
        else:
            query = query.order_by(User.username.desc())
    elif sort_by == 'loan':
        if sort_order == 'asc':
            query = query.order_by(Loan.loan_name.asc())
        else:
            query = query.order_by(Loan.loan_name.desc())
    elif sort_by == 'amount':
        if sort_order == 'asc':
            query = query.order_by(Payment.amount.asc())
        else:
            query = query.order_by(Payment.amount.desc())
    else:  # payment_date
        if sort_order == 'asc':
            query = query.order_by(Payment.payment_date.asc())
        else:
            query = query.order_by(Payment.payment_date.desc())
    
    payments = query.all()
    
    # Calculate cashback for each payment
    payment_cashback_map = {}
    payment_cashback_recipients_map = {}
    for payment, loan, user in payments:
        payment_id = payment.id
        cashback_total = get_payment_cashback_total(payment_id, instance_name)
        payment_cashback_map[payment_id] = cashback_total
        
        # Get cashback recipients for this payment
        session = db_manager.get_session_for_instance(instance)
        cashback_transactions = session.query(CashbackTransaction).filter_by(
            related_payment_id=payment_id
        ).all()
        recipients = []
        for txn in cashback_transactions:
            if txn.to_user:
                recipients.append({
                    'username': txn.to_user.username,
                    'points': txn.points
                })
        payment_cashback_recipients_map[payment_id] = recipients
    
    # Calculate total interest paid (verified payments only)
    total_interest_paid = get_payment_query().filter_by(status='verified').with_entities(db.func.sum(Payment.interest_amount)).scalar() or 0
    
    # Calculate filtered totals
    filtered_principal = sum(payment.principal_amount for payment, loan, user in payments)
    filtered_interest = sum(payment.interest_amount for payment, loan, user in payments)
    filtered_total = sum(payment.amount for payment, loan, user in payments)
    
    # Get unique customers and loans for filters
    customers = [user.username for user in get_user_query().filter_by(is_admin=False).all()]
    loans = [loan.loan_name for loan in get_loan_query().all()]
    
    # Handle Excel export
    if request.args.get('export') == 'excel':
        from backup_multi import MultiInstanceBackupManager
        backup_manager = MultiInstanceBackupManager(app)
        return backup_manager.export_to_excel(instance_name)
    
    return render_template('admin/payments.html', 
                         payments=payments,
                         payment_cashback_map=payment_cashback_map,
                         payment_cashback_recipients_map=payment_cashback_recipients_map,
                         total_interest_paid=total_interest_paid,
                         filtered_principal=filtered_principal,
                         filtered_interest=filtered_interest,
                         filtered_total=filtered_total,
                         customers=customers,
                         loans=loans,
                         customer_filter=customer_filter,
                         loan_filter=loan_filter,
                         status_filter=status_filter,
                         payment_method=payment_method,
                         sort_by=sort_by,
                         sort_order=sort_order,
                         instance_name=instance_name)

@app.route('/<instance_name>/admin/payments/pending')
@login_required
def admin_pending_payments(instance_name):
    """Admin view all pending payments for approval"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    # Get the current instance session
    instance = get_current_instance_from_g()
    session = db_manager.get_session_for_instance(instance)
    
    # Get all pending payments
    pending_payments = session.query(Payment, Loan, User).join(
        Loan, Payment.loan_id == Loan.id
    ).join(
        User, Loan.customer_id == User.id
    ).filter(
        Payment.status == 'pending'
    ).order_by(
        Payment.payment_date.desc()
    ).all()
    
    return render_template('admin/pending_payments.html',
                         pending_payments=pending_payments,
                         instance_name=instance_name)

# Admin Add Payment route
@app.route('/<instance_name>/admin/add-payment', methods=['GET', 'POST'])
@app.route('/<instance_name>/admin/add-payment/<int:loan_id>', methods=['GET', 'POST'])
@login_required
def admin_add_payment(instance_name, loan_id=None):
    """Admin add payment page for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    if request.method == 'POST':
        loan_id = request.form['loan_id']
        amount = Decimal(request.form['amount'])
        payment_method = request.form['payment_method']
        transaction_id = request.form.get('transaction_id', '')
        payment_date_str = request.form.get('payment_date')
        principal_only = request.form.get('principal_only') == 'on'  # Checkbox value
        
        # Parse payment date
        if payment_date_str:
            try:
                payment_date = datetime.strptime(payment_date_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Invalid date format', 'error')
                return render_template('admin/add_payment.html', 
                                     loans=get_loan_query().all(),
                                     selected_loan_id=loan_id,
                                     instance_name=instance_name)
        else:
            payment_date = datetime.utcnow()
        
        loan = get_loan_query().get(loan_id)
        if not loan:
            flash('Loan not found', 'error')
            return redirect(url_for('admin_payments', instance_name=instance_name))
        
        # Handle principal-only payment
        if principal_only:
            # Principal-only payment: all goes to principal, no interest
            interest_amount = Decimal('0')
            principal_amount = amount
            
            # Create payment directly (bypass process_payment)
            payment = Payment(
                loan_id=loan.id,
                amount=amount,
                payment_date=payment_date,
                payment_type='principal',  # Payment type for principal-only
                interest_amount=interest_amount,
                principal_amount=principal_amount,
                transaction_id=transaction_id,
                payment_method=payment_method,
                status='pending'
            )
            
            add_to_current_instance(payment)
            commit_current_instance()
            
            # Send notification to admins for approval
            try:
                from app_notifications import send_approval_notification
                send_approval_notification(
                    instance_name=instance_name,
                    approval_type='payment',
                    item_id=payment.id,
                    item_details={
                        'loan_name': loan.loan_name,
                        'customer_name': loan.customer.username,
                        'amount': f'{amount:,.2f}',
                        'payment_date': payment_date.strftime('%Y-%m-%d %H:%M'),
                        'payment_method': payment_method
                    }
                )
            except Exception as e:
                print(f"Error sending payment notification: {e}")
                # Don't fail the payment creation if notification fails
            
            flash(f'Principal-only payment of ₹{amount:,.2f} added successfully. It will be verified by admin.', 'success')
            return redirect(url_for('admin_payments', instance_name=instance_name))
        
        # Regular payment processing (existing logic)
        try:
            payment = process_payment(
                loan=loan,
                payment_amount=amount,
                payment_date=payment_date,
                transaction_id=transaction_id,
                payment_method=payment_method,
                proof_filename=None
            )
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(url_for('admin_add_payment', instance_name=instance_name, loan_id=loan_id))
        
        # Send notification to admins for approval
        try:
            from app_notifications import send_approval_notification
            send_approval_notification(
                instance_name=instance_name,
                approval_type='payment',
                item_id=payment.id,
                item_details={
                    'loan_name': loan.loan_name,
                    'customer_name': loan.customer.username,
                    'amount': f'{amount:,.2f}',
                    'payment_date': payment_date.strftime('%Y-%m-%d %H:%M'),
                    'payment_method': payment_method
                }
            )
        except Exception as e:
            print(f"Error sending payment notification: {e}")
            # Don't fail the payment creation if notification fails
        
        flash('Payment added successfully. It will be verified by admin.', 'success')
        return redirect(url_for('admin_payments', instance_name=instance_name))
    
    loans = get_loan_query().all()
    return render_template('admin/add_payment.html', 
                         loans=loans,
                         selected_loan_id=loan_id,
                         instance_name=instance_name)

# Admin Backup route
# Admin Toggle Moderator Status route
@app.route('/<instance_name>/admin/toggle-moderator/<int:user_id>', methods=['POST'])
@login_required
def admin_toggle_moderator(instance_name, user_id):
    """Toggle moderator status for a user"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    user = get_user_query().filter_by(id=user_id).first()
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    if user.is_admin:
        flash('Cannot change moderator status for admin users', 'error')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    # Toggle moderator status
    user.is_moderator = not user.is_moderator
    commit_current_instance()
    
    if user.is_moderator:
        flash(f'{user.username} is now a moderator', 'success')
    else:
        flash(f'{user.username} moderator status removed', 'success')
    
    return redirect(url_for('admin_users', instance_name=instance_name))


# Admin Edit Payment route
@app.route('/<instance_name>/admin/edit-payment/<int:payment_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_payment(instance_name, payment_id):
    """Admin edit payment page for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    payment = get_payment_query().filter_by(id=payment_id).first()
    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('admin_payments', instance_name=instance_name))
    
    loan = payment.loan  # Get the loan associated with this payment
    
    if request.method == 'POST':
        # Store old values for loan recalculation
        old_status = payment.status
        old_principal_amount = payment.principal_amount
        old_interest_amount = payment.interest_amount
        
        # Get form data
        new_amount = Decimal(request.form.get('amount', 0))
        new_interest_amount = Decimal(request.form.get('interest_amount', 0))
        new_principal_amount = Decimal(request.form.get('principal_amount', 0))
        new_transaction_id = request.form.get('transaction_id', '')
        new_payment_method = request.form.get('payment_method', '')
        new_status = request.form.get('status', 'pending')
        
        # Validate amounts
        if new_amount != new_interest_amount + new_principal_amount:
            flash('Total amount must equal interest + principal amount', 'error')
            return render_template('admin/edit_payment.html', 
                                 payment=payment,
                                 loan=loan,
                                 instance_name=instance_name)
        
        # Update payment
        payment.amount = new_amount
        payment.interest_amount = new_interest_amount
        payment.principal_amount = new_principal_amount
        payment.transaction_id = new_transaction_id
        payment.payment_method = new_payment_method
        payment.status = new_status
        
        # Parse payment date if provided
        payment_date_str = request.form.get('payment_date')
        if payment_date_str:
            try:
                payment.payment_date = datetime.strptime(payment_date_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Invalid date format', 'error')
                return render_template('admin/edit_payment.html', 
                                     payment=payment,
                                     loan=loan,
                                     instance_name=instance_name)
        
        # Handle status changes and balance adjustments
        if new_status == 'verified' and old_status == 'pending':
            # Payment is being verified - reduce balance
            loan.remaining_principal -= new_principal_amount
        elif new_status == 'pending' and old_status == 'verified':
            # Payment is being unverified - add back to balance
            loan.remaining_principal += old_principal_amount
        elif new_status == 'verified' and old_status == 'verified':
            # Payment was already verified, adjust for amount changes
            principal_difference = new_principal_amount - old_principal_amount
            loan.remaining_principal -= principal_difference
        
        # Ensure remaining principal doesn't go negative
        loan.remaining_principal = max(Decimal('0'), loan.remaining_principal)
        
        commit_current_instance()
        flash('Payment updated successfully', 'success')
        return redirect(url_for('admin_payments', instance_name=instance_name))
    
    # GET request - show edit form
    return render_template('admin/edit_payment.html', 
                         payment=payment,
                         loan=loan,
                         instance_name=instance_name)


# Admin Delete Payment route
@app.route('/<instance_name>/admin/delete-payment/<int:payment_id>', methods=['POST'])
@login_required
def admin_delete_payment(instance_name, payment_id):
    """Admin delete payment for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    payment = get_payment_query().filter_by(id=payment_id).first()
    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('admin_payments', instance_name=instance_name))
    
    loan = payment.loan
    
    # Store payment details before deletion
    payment_amount = payment.amount
    payment_status = payment.status
    payment_principal = payment.principal_amount
    loan_id = loan.id
    
    try:
        # If payment was verified, we need to adjust the loan balance
        if payment_status == 'verified':
            # Add back the principal amount to remaining_principal
            loan.remaining_principal += payment_principal
            # Ensure remaining_principal doesn't exceed original principal
            if loan.remaining_principal > loan.principal_amount:
                loan.remaining_principal = loan.principal_amount
        
        # Delete the payment
        get_payment_query().filter_by(id=payment_id).delete()
        commit_current_instance()
        
        # Log admin action
        try:
            logging_mgr = get_logging_manager(instance_name)
            metrics_mgr = get_metrics_manager(instance_name)
            logging_mgr.log_admin_action('delete_payment', 'payment', payment_id, 
                                       username=current_user.username,
                                       details={'payment_amount': str(payment_amount), 'loan_id': loan_id})
            metrics_mgr.record_admin_action('delete_payment', current_user.username)
        except Exception as log_error:
            print(f"[ERROR] Failed to log admin action: {log_error}")
        
        flash(f'Payment of ₹{payment_amount:,.2f} deleted successfully', 'success')
        return redirect(url_for('admin_payments', instance_name=instance_name))
        
    except Exception as e:
        flash(f'Error deleting payment: {str(e)}', 'error')
        return redirect(url_for('admin_payments', instance_name=instance_name))


# Tracker moderator assignment routes moved to app_trackers.py
# Loan routes (admin_edit_loan, admin_view_loan, admin_close_loan, admin_delete_loan, admin_split_loan, admin_assign_payment_to_split, admin_assign_moderator_to_loan, admin_unassign_moderator_from_loan, admin_loan_excel, admin_delete_payment) moved to app_loans.py
# Customer loan routes (customer_all_loans, customer_loan_detail, customer_edit_notes) moved to app_loans.py

# Admin Create Backup route
# Customer routes
@app.route('/<instance_name>/customer')
@login_required
def customer_dashboard(instance_name):
    """Customer dashboard for specific instance"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard', instance_name=instance_name))
    
    # Get customer's active loans only (exclude closed)
    loans = get_loan_query().filter_by(
        customer_id=current_user.id, 
        is_active=True,
        status='active'
    ).all()
    
    loan_data = []
    for loan in loans:
        daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
        monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
        interest_data = calculate_accumulated_interest(loan)
        accumulated_interest_daily = interest_data['daily']
        accumulated_interest_monthly = interest_data['monthly']
        
        # Calculate pending payments for this specific loan
        pending_payments = get_payment_query().filter_by(loan_id=loan.id, status='pending').all()
        pending_principal = sum(payment.principal_amount for payment in pending_payments)
        pending_interest = sum(payment.interest_amount for payment in pending_payments)
        pending_total = sum(payment.amount for payment in pending_payments)
        
        # Calculate verified payments for this specific loan
        verified_payments = get_payment_query().filter_by(loan_id=loan.id, status='verified').all()
        verified_principal = sum(payment.principal_amount for payment in verified_payments)
        verified_interest = sum(payment.interest_amount for payment in verified_payments)
        
        loan_data.append({
            'loan': loan,
            'daily_interest': daily_interest,
            'monthly_interest': monthly_interest,
            'accumulated_interest_daily': accumulated_interest_daily,
            'accumulated_interest_monthly': accumulated_interest_monthly,
            'interest_data': interest_data,
            'pending_principal': pending_principal,
            'pending_interest': pending_interest,
            'pending_total': pending_total,
            'verified_principal': verified_principal,
            'verified_interest': verified_interest
        })
    
    # Check if user has daily trackers (active and not closed by user)
    daily_trackers = get_daily_tracker_query().filter_by(
        user_id=current_user.id, 
        is_active=True,
        is_closed_by_user=False
    ).all()
    
    # Check if user has moderator assignments (assigned loans or trackers)
    assigned_loans_count = len(current_user.assigned_loans.filter_by(is_active=True).all())
    assigned_trackers_count = len(current_user.assigned_trackers.filter_by(is_active=True).all())
    has_moderator_assignments = assigned_loans_count > 0 or assigned_trackers_count > 0
    
    # Get cashback balance
    cashback_balance = get_user_cashback_balance(current_user.id, instance_name)
    
    return render_template('customer/dashboard.html',
                         loan_data=loan_data,
                         daily_trackers=daily_trackers,
                         has_moderator_assignments=has_moderator_assignments,
                         assigned_items_count=assigned_loans_count + assigned_trackers_count,
                         cashback_balance=cashback_balance,
                         instance_name=instance_name)

# Tracker routes moved to app_trackers.py - removing duplicates

# Password Management Routes

# Change Password route
@app.route('/<instance_name>/change-password', methods=['GET', 'POST'])
@login_required
def change_password(instance_name):
    """Change password for current user"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Verify current password
        if not check_password_hash(current_user.password_hash, current_password):
            flash('Current password is incorrect', 'error')
            return render_template('change_password.html', instance_name=instance_name)
        
        # Validate new password
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_password.html', instance_name=instance_name)
        
        if len(new_password) < 1:
            flash('Password cannot be empty', 'error')
            return render_template('change_password.html', instance_name=instance_name)
        
        # Update password
        current_user.password_hash = generate_password_hash(new_password)
        commit_current_instance()
        
        flash('Password changed successfully', 'success')
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'customer_dashboard', instance_name=instance_name))
    
    return render_template('change_password.html', instance_name=instance_name)

# User Settings route
@app.route('/<instance_name>/settings', methods=['GET', 'POST'])
@login_required
def user_settings(instance_name):
    """User settings page for notification preferences and account settings"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    session_db = db_manager.get_session_for_instance(instance_name)
    
    # Get or create notification preference for the user
    notification_pref = session_db.query(NotificationPreference).filter_by(
        user_id=current_user.id
    ).first()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'notifications':
            # Update email address if provided
            email = request.form.get('email', '').strip()
            email_enabled = request.form.get('email_enabled') == '1'
            
            if email:
                current_user.email = email
            
            # Get notification type preferences
            preferences = {}
            if current_user.is_admin:
                preferences['payment_approvals'] = request.form.get('payment_approvals') == '1'
                preferences['tracker_approvals'] = request.form.get('tracker_approvals') == '1'
            else:
                preferences['payment_status'] = request.form.get('payment_status') == '1'
                preferences['tracker_status'] = request.form.get('tracker_status') == '1'
            
            # Create or update notification preference
            if notification_pref:
                notification_pref.enabled = email_enabled
                notification_pref.preferences = preferences
                notification_pref.updated_at = datetime.utcnow()
            else:
                notification_pref = NotificationPreference(
                    user_id=current_user.id,
                    channel='email',
                    enabled=email_enabled,
                    preferences=preferences
                )
                session_db.add(notification_pref)
            
            commit_current_instance()
            flash('Notification settings saved successfully', 'success')
            return redirect(url_for('user_settings', instance_name=instance_name))
    
    # Parse preferences for template
    preferences = notification_pref.preferences if notification_pref and notification_pref.preferences else {}
    
    return render_template('user_settings.html', 
                         instance_name=instance_name,
                         notification_pref=notification_pref,
                         preferences=preferences)

# Forgot Password route
@app.route('/<instance_name>/forgot-password', methods=['GET', 'POST'])
def forgot_password(instance_name):
    """Forgot password - send reset email"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if request.method == 'POST':
        email = request.form['email']
        
        # Find user by email
        user = get_user_query().filter_by(email=email).first()
        
        if user:
            # Generate reset token
            reset_token = secrets.token_urlsafe(32)
            user.reset_token = reset_token
            user.reset_token_expires = datetime.utcnow() + timedelta(hours=1)  # Token expires in 1 hour
            commit_current_instance()
            
            # Send reset email (simplified - in production, use proper email service)
            try:
                send_password_reset_email(user.email, reset_token, instance_name)
                flash('Password reset instructions have been sent to your email', 'success')
            except Exception as e:
                flash('Error sending email. Please contact administrator.', 'error')
        else:
            flash('No account found with that email address', 'error')
    
    return render_template('forgot_password.html', instance_name=instance_name)

# Reset Password route
@app.route('/<instance_name>/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(instance_name, token):
    """Reset password with token"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    # Find user by token
    user = get_user_query().filter_by(reset_token=token).first()
    
    if not user or not user.reset_token_expires or user.reset_token_expires < datetime.utcnow():
        flash('Invalid or expired reset token', 'error')
        return redirect(url_for('login', instance_name=instance_name))
    
    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Validate passwords
        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('reset_password.html', token=token, instance_name=instance_name)
        
        if len(new_password) < 1:
            flash('Password cannot be empty', 'error')
            return render_template('reset_password.html', token=token, instance_name=instance_name)
        
        # Update password and clear reset token
        user.password_hash = generate_password_hash(new_password)
        user.reset_token = None
        user.reset_token_expires = None
        commit_current_instance()
        
        flash('Password reset successfully. Please log in with your new password.', 'success')
        return redirect(url_for('login', instance_name=instance_name))
    
    return render_template('reset_password.html', token=token, instance_name=instance_name)

# Admin Reset User Password route
@app.route('/<instance_name>/admin/reset-user-password/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_reset_user_password(instance_name, user_id):
    """Admin reset any user's password"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    user = get_user_query().filter_by(id=user_id).first()
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Validate passwords
        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('admin/reset_user_password.html', user=user, instance_name=instance_name)
        
        if len(new_password) < 1:
            flash('Password cannot be empty', 'error')
            return render_template('admin/reset_user_password.html', user=user, instance_name=instance_name)
        
        # Update password
        user.password_hash = generate_password_hash(new_password)
        commit_current_instance()
        
        flash(f'Password reset successfully for user {user.username}', 'success')
        return redirect(url_for('admin_users', instance_name=instance_name))
    
    return render_template('admin/reset_user_password.html', user=user, instance_name=instance_name)

# ============================================================================
# DAILY TRACKER ROUTES - Moved to app_trackers.py

# Email helper function
def send_password_reset_email(email, token, instance_name):
    """Send password reset email (simplified implementation)"""
    # In production, use proper email service like SendGrid, AWS SES, etc.
    # For now, we'll just print the reset link to console
    reset_url = f"http://127.0.0.1:8080/{instance_name}/reset-password/{token}"
    
    print(f"\n=== PASSWORD RESET EMAIL ===")
    print(f"To: {email}")
    print(f"Subject: Password Reset Request")
    print(f"Reset Link: {reset_url}")
    print(f"=============================\n")
    
    # In production, replace this with actual email sending:
    # msg = MIMEMultipart()
    # msg['From'] = "noreply@lendingapp.com"
    # msg['To'] = email
    # msg['Subject'] = "Password Reset Request"
    # 
    # body = f"""
    # You requested a password reset for your account.
    # 
    # Click the link below to reset your password:
    # {reset_url}
    # 
    # This link will expire in 1 hour.
    # 
    # If you didn't request this, please ignore this email.
    # """
    # 
    # msg.attach(MIMEText(body, 'plain'))
    # 
    # server = smtplib.SMTP('smtp.gmail.com', 587)
    # server.starttls()
    # server.login("your-email@gmail.com", "your-password")
    # text = msg.as_string()
    # server.sendmail("noreply@lendingapp.com", email, text)
    # server.quit()


# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================
    """Admin create new daily tracker"""
    if instance_name not in VALID_INSTANCES:
        return redirect('/')
    
    if not current_user.is_admin:
        flash('Access denied', 'error')
        return redirect(url_for('customer_dashboard', instance_name=instance_name))
    
    if request.method == 'POST':
        try:
            user_id = int(request.form['user_id'])
            tracker_name = request.form['tracker_name']
            tracker_type = request.form['tracker_type']
            investment = Decimal(request.form['investment'])
            scheme_period = int(request.form['scheme_period'])
            per_day_payment = Decimal(request.form['per_day_payment'])
            start_date_str = request.form['start_date']
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            
            # Validate tracker type
            if tracker_type not in TRACKER_TYPES:
                flash('Invalid tracker type', 'error')
                users = get_user_query().filter_by(is_admin=False).all()
                return render_template('admin/create_daily_tracker.html',
                                     users=users,
                                     instance_name=instance_name,
                                     tracker_types=TRACKER_TYPES)
            
            # Get user
            user = get_user_query().filter_by(id=user_id).first()
            if not user:
                flash('User not found', 'error')
                users = get_user_query().filter_by(is_admin=False).all()
                return render_template('admin/create_daily_tracker.html',
                                     users=users,
                                     instance_name=instance_name,
                                     tracker_types=TRACKER_TYPES)
            
            # Create Excel file
            filename = create_tracker_file(
                instance_name,
                user.username,
                tracker_name,
                tracker_type,
                investment,
                scheme_period,
                start_date,
                per_day_payment
            )
            
            # Create database entry
            tracker = DailyTracker(
                user_id=user_id,
                tracker_name=tracker_name,
                tracker_type=tracker_type,
                investment=investment,
                scheme_period=scheme_period,
                per_day_payment=per_day_payment,
                start_date=start_date,
                filename=filename
            )
            
            add_to_current_instance(tracker)
            commit_current_instance()
            
            flash(f'Daily tracker created successfully for {user.username}', 'success')
            return redirect(url_for('admin_daily_trackers', instance_name=instance_name))
            
        except Exception as e:
            flash(f'Error creating daily tracker: {str(e)}', 'error')
            print(f"Error creating daily tracker: {e}")
    
    # GET request
    users = get_user_query().filter_by(is_admin=False).all()
    return render_template('admin/create_daily_tracker.html',
                         users=users,
                         instance_name=instance_name,
                         tracker_types=TRACKER_TYPES)


# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================


# Email helper function
def send_password_reset_email(email, token, instance_name):
    """Send password reset email (simplified implementation)"""
    # In production, use proper email service like SendGrid, AWS SES, etc.
    # For now, we'll just print the reset link to console
    reset_url = f"http://127.0.0.1:8080/{instance_name}/reset-password/{token}"
    
    print(f"\n=== PASSWORD RESET EMAIL ===")
    print(f"To: {email}")
    print(f"Subject: Password Reset Request")
    print(f"Reset Link: {reset_url}")
    print(f"=============================\n")
    
    # In production, replace this with actual email sending:
    # msg = MIMEMultipart()
    # msg['From'] = "noreply@lendingapp.com"
    # msg['To'] = email
    # msg['Subject'] = "Password Reset Request"
    # 
    # body = f"""
    # You requested a password reset for your account.
    # 
    # Click the link below to reset your password:
    # {reset_url}
    # 
    # This link will expire in 1 hour.
    # 
    # If you didn't request this, please ignore this email.
    # """
    # 
    # msg.attach(MIMEText(body, 'plain'))
    # 
    # server = smtplib.SMTP('smtp.gmail.com', 587)
    # server.starttls()
    # server.login("your-email@gmail.com", "your-password")
    # text = msg.as_string()
    # server.sendmail("noreply@lendingapp.com", email, text)
    # server.quit()


# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================


# Initialize the app only when run directly
# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================


# Initialize the app only when run directly
# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================


# ============================================================================
# MODERATOR ROUTES - Moved to app_moderator.py
# ============================================================================


# Initialize the app only when run directly
if __name__ == '__main__':
    init_app()
    app.run(debug=True, port=8080)
