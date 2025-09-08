from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import os
import uuid

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///lending_app.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# File upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships are defined in the Loan model

class InterestRate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rate = db.Column(db.Numeric(10, 4), nullable=False)  # Annual interest rate as decimal
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

class Loan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    loan_name = db.Column(db.String(100), nullable=False)  # e.g., "Personal Loan", "Business Loan"
    principal_amount = db.Column(db.Numeric(15, 2), nullable=False)
    remaining_principal = db.Column(db.Numeric(15, 2), nullable=False)
    interest_rate = db.Column(db.Numeric(5, 4), nullable=False)  # Store as decimal (e.g., 0.21 for 21%)
    payment_frequency = db.Column(db.String(20), nullable=False)  # 'daily' or 'monthly'
    loan_type = db.Column(db.String(20), nullable=False, default='regular')  # 'regular' or 'interest_only'
    admin_notes = db.Column(db.Text, nullable=True)  # Private admin notes (not visible to customers)
    customer_notes = db.Column(db.Text, nullable=True)  # Customer notes (visible to both admin and customer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationships
    customer = db.relationship('User', backref='loans')
    payments = db.relationship('Payment', backref='loan', lazy=True, order_by='Payment.payment_date.desc()')

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    amount = db.Column(db.Numeric(15, 2), nullable=False)
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)
    payment_type = db.Column(db.String(20), nullable=False)  # 'interest', 'principal', 'both'
    interest_amount = db.Column(db.Numeric(15, 2), default=0)
    principal_amount = db.Column(db.Numeric(15, 2), default=0)
    transaction_id = db.Column(db.String(100), nullable=True)  # UPI transaction ID
    payment_method = db.Column(db.String(20), nullable=True)  # 'gpay', 'upi', 'phonepay', 'bank_transfer'
    proof_filename = db.Column(db.String(255), nullable=True)  # Uploaded proof file name
    status = db.Column(db.String(20), default='pending')  # 'pending', 'verified', 'rejected'

class PendingInterest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    amount = db.Column(db.Numeric(15, 2), nullable=False)
    due_date = db.Column(db.Date, nullable=False)
    month_year = db.Column(db.String(7), nullable=False)  # Format: "2024-01"
    is_paid = db.Column(db.Boolean, default=False)
    
    # Relationship
    loan = db.relationship('Loan', backref='pending_interests')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Helper functions for file upload
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generate_unique_filename(original_filename):
    """Generate a unique filename to prevent conflicts"""
    ext = original_filename.rsplit('.', 1)[1].lower()
    unique_id = str(uuid.uuid4())
    return f"{unique_id}.{ext}"

# Helper Functions
def calculate_daily_interest(principal, annual_rate):
    """Calculate daily interest amount"""
    daily_rate = Decimal(str(annual_rate)) / Decimal('365')
    return (Decimal(str(principal)) * daily_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def calculate_monthly_interest(principal, annual_rate):
    """Calculate monthly interest amount"""
    monthly_rate = Decimal(str(annual_rate)) / Decimal('12')
    return (Decimal(str(principal)) * monthly_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def calculate_interest_for_period(principal, annual_rate, start_date, end_date):
    """Calculate interest for a specific time period"""
    from datetime import datetime
    
    if isinstance(start_date, str):
        start_date = datetime.fromisoformat(start_date)
    if isinstance(end_date, str):
        end_date = datetime.fromisoformat(end_date)
    
    # Calculate days between dates
    days_elapsed = (end_date - start_date).days
    
    if days_elapsed <= 0:
        return Decimal('0')
    
    # Calculate daily rate
    daily_rate = Decimal(str(annual_rate)) / Decimal('365')
    
    # Calculate interest for the period
    interest = Decimal(str(principal)) * daily_rate * Decimal(str(days_elapsed))
    
    return interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def calculate_accumulated_interest(loan, as_of_date=None):
    """Calculate total accumulated interest for a loan from creation to a specific date"""
    from datetime import datetime
    
    if as_of_date is None:
        as_of_date = datetime.utcnow()
    
    if isinstance(as_of_date, str):
        as_of_date = datetime.fromisoformat(as_of_date)
    
    # Calculate interest from loan creation to the specified date
    total_interest = calculate_interest_for_period(
        loan.remaining_principal, 
        loan.interest_rate, 
        loan.created_at, 
        as_of_date
    )
    
    # Subtract any verified interest payments
    verified_interest_paid = db.session.query(db.func.sum(Payment.interest_amount)).filter(
        Payment.loan_id == loan.id,
        Payment.status == 'verified'
    ).scalar() or Decimal('0')
    
    accumulated_interest = total_interest - verified_interest_paid
    
    return max(accumulated_interest, Decimal('0'))


def process_payment(loan, payment_amount, payment_date=None, transaction_id=None, payment_method=None, proof_filename=None):
    """Process a payment and update loan accordingly"""
    if payment_date is None:
        payment_date = datetime.utcnow()
    
    payment_amount = Decimal(str(payment_amount))
    
    # For interest-only loans, customers can only pay interest
    if loan.loan_type == 'interest_only':
        # Calculate accumulated interest from loan creation to payment date
        accumulated_interest = calculate_accumulated_interest(loan, payment_date)
        
        # Get total pending interest from all pending payments
        pending_interest = db.session.query(db.func.sum(Payment.interest_amount)).filter(
            Payment.loan_id == loan.id,
            Payment.status == 'pending'
        ).scalar() or Decimal('0')
        
        total_pending_interest = accumulated_interest + pending_interest
        
        # Limit payment to pending interest only
        if payment_amount > total_pending_interest:
            raise ValueError(f"Payment amount (₹{payment_amount}) exceeds pending interest (₹{total_pending_interest}). For interest-only loans, you can only pay up to the pending interest amount.")
        
        interest_amount = payment_amount
        principal_amount = Decimal('0')
        payment_type = 'interest'
        
    else:  # Regular loans
        # Calculate interest based on payment frequency
        if loan.payment_frequency == 'daily':
            # Calculate daily interest
            daily_rate = loan.interest_rate / Decimal('365')
            interest_amount = loan.remaining_principal * daily_rate
        else:  # monthly
            # Calculate monthly interest
            monthly_rate = loan.interest_rate / Decimal('12')
            interest_amount = loan.remaining_principal * monthly_rate
        
        # Round to 2 decimal places
        interest_amount = interest_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # For regular loans, allow payment of interest + principal
        interest_amount = min(payment_amount, interest_amount)
        principal_amount = payment_amount - interest_amount
        payment_type = 'both' if principal_amount > 0 else 'interest'
    
    # All payments require admin verification
    payment_status = 'pending'
    
    # Only reduce remaining principal for verified payments (none initially)
    # Balance will be reduced only when admin verifies the payment
    
    # Create payment record
    payment = Payment(
        loan_id=loan.id,
        amount=payment_amount,
        payment_date=payment_date,
        payment_type=payment_type,
        interest_amount=interest_amount,
        principal_amount=principal_amount,
        transaction_id=transaction_id,
        payment_method=payment_method,
        proof_filename=proof_filename,
        status=payment_status
    )
    
    db.session.add(payment)
    db.session.commit()
    
    return payment

def verify_payment(payment_id):
    """Verify a pending payment and reduce loan balance"""
    payment = Payment.query.get_or_404(payment_id)
    loan = Loan.query.get(payment.loan_id)
    
    if payment.status == 'pending':
        # For interest-only loans, principal never changes
        if loan.loan_type != 'interest_only':
            # Reduce remaining principal for regular loans
            loan.remaining_principal -= payment.principal_amount
        
        # Update payment status
        payment.status = 'verified'
        
        db.session.commit()
        return True
    return False

def export_payments_to_excel(payments):
    """Export payments data to Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import make_response
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    # Define headers
    headers = [
        'Payment ID', 'Date', 'Customer', 'Email', 'Loan Name', 'Interest Rate (%)',
        'Total Amount (₹)', 'Interest Amount (₹)', 'Principal Amount (₹)',
        'Payment Method', 'Transaction ID', 'Status', 'Proof File'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, (payment, loan, user) in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment.id)
        ws.cell(row=row, column=2, value=payment.payment_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=user.username)
        ws.cell(row=row, column=4, value=user.email)
        ws.cell(row=row, column=5, value=loan.loan_name)
        ws.cell(row=row, column=6, value=float(loan.interest_rate * 100))
        ws.cell(row=row, column=7, value=float(payment.amount))
        ws.cell(row=row, column=8, value=float(payment.interest_amount))
        ws.cell(row=row, column=9, value=float(payment.principal_amount))
        ws.cell(row=row, column=10, value=payment.payment_method.title() if payment.payment_method else '')
        ws.cell(row=row, column=11, value=payment.transaction_id or '')
        ws.cell(row=row, column=12, value=payment.status.title())
        ws.cell(row=row, column=13, value=payment.proof_filename or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=payments_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('customer_dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            flash('Invalid username or password')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form.get('email', '').strip()  # Get email, default to empty string
        password = request.form['password']
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return render_template('register.html')
        
        # Only check email uniqueness if email is provided
        if email and User.query.filter_by(email=email).first():
            flash('Email already exists')
            return render_template('register.html')
        
        user = User(
            username=username,
            email=email if email else None,  # Store as None if empty
            password_hash=generate_password_hash(password)
        )
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# Admin Routes
@app.route('/admin')
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    # Get statistics
    total_loans = Loan.query.filter_by(is_active=True).count()
    total_principal = db.session.query(db.func.sum(Loan.remaining_principal)).filter_by(is_active=True).scalar() or 0
    
    # Calculate total interest earned (only interest payments, not principal)
    total_interest_earned = db.session.query(db.func.sum(Payment.interest_amount)).scalar() or 0
    
    return render_template('admin/dashboard.html', 
                         total_loans=total_loans,
                         total_principal=total_principal,
                         total_interest_earned=total_interest_earned)

@app.route('/admin/interest-rate')
@login_required
def admin_interest_rate():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    # This page is now informational only since each loan has its own rate
    return render_template('admin/interest_rate.html')

@app.route('/admin/loans')
@login_required
def admin_loans():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    # Get filter parameters
    loan_type = request.args.get('loan_type', '')
    frequency = request.args.get('frequency', '')
    customer_id = request.args.get('customer', '')
    status = request.args.get('status', '')
    loan_name = request.args.get('loan_name', '')
    min_rate = request.args.get('min_rate', '')
    max_rate = request.args.get('max_rate', '')
    sort_by = request.args.get('sort', 'created')
    sort_order = request.args.get('order', 'desc')
    
    # Start with base query
    query = Loan.query.filter_by(is_active=True)
    
    # Apply filters
    if loan_type:
        query = query.filter_by(loan_type=loan_type)
    
    if frequency:
        query = query.filter_by(payment_frequency=frequency)
    
    if customer_id:
        query = query.filter_by(customer_id=customer_id)
    
    if status == 'paid_off':
        query = query.filter(Loan.remaining_principal == 0)
    elif status == 'active':
        query = query.filter(Loan.remaining_principal > 0)
    
    if loan_name:
        query = query.filter(Loan.loan_name.ilike(f'%{loan_name}%'))
    
    if min_rate:
        try:
            min_rate_decimal = Decimal(min_rate) / 100  # Convert percentage to decimal
            query = query.filter(Loan.interest_rate >= min_rate_decimal)
        except (ValueError, InvalidOperation):
            pass  # Ignore invalid input
    
    if max_rate:
        try:
            max_rate_decimal = Decimal(max_rate) / 100  # Convert percentage to decimal
            query = query.filter(Loan.interest_rate <= max_rate_decimal)
        except (ValueError, InvalidOperation):
            pass  # Ignore invalid input
    
    # Apply sorting
    if sort_by == 'id':
        order_column = Loan.id
    elif sort_by == 'loan_name':
        order_column = Loan.loan_name
    elif sort_by == 'customer':
        order_column = User.username
        query = query.join(User)
    elif sort_by == 'loan_type':
        order_column = Loan.loan_type
    elif sort_by == 'principal':
        order_column = Loan.principal_amount
    elif sort_by == 'remaining':
        order_column = Loan.remaining_principal
    elif sort_by == 'interest_rate':
        order_column = Loan.interest_rate
    elif sort_by == 'frequency':
        order_column = Loan.payment_frequency
    elif sort_by == 'created':
        order_column = Loan.created_at
    else:
        order_column = Loan.created_at
    
    # Apply sort order
    if sort_order == 'asc':
        query = query.order_by(order_column.asc())
    else:
        query = query.order_by(order_column.desc())
    
    loans = query.all()
    
    # Get all customers for filter dropdown
    customers = User.query.filter_by(is_admin=False).all()
    
    return render_template('admin/loans.html', loans=loans, customers=customers)

@app.route('/admin/create-loan', methods=['GET', 'POST'])
@login_required
def admin_create_loan():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    if request.method == 'POST':
        customer_id = request.form['customer_id']
        loan_name = request.form['loan_name']
        principal = Decimal(request.form['principal'])
        payment_frequency = request.form['payment_frequency']
        loan_type = request.form['loan_type']
        interest_rate = Decimal(request.form['interest_rate']) / 100  # Convert percentage to decimal
        admin_notes = request.form.get('admin_notes', '').strip()  # Get admin notes
        customer_notes = request.form.get('customer_notes', '').strip()  # Get customer notes
        
        # Handle custom creation date
        custom_created_at = None
        loan_created_date = request.form.get('loan_created_date', '').strip()
        loan_created_time = request.form.get('loan_created_time', '').strip()
        
        if loan_created_date:
            try:
                from datetime import datetime
                if loan_created_time:
                    # Combine date and time
                    custom_created_at = datetime.strptime(f"{loan_created_date} {loan_created_time}", "%Y-%m-%d %H:%M")
                else:
                    # Use date only, default to 12:00 PM
                    custom_created_at = datetime.strptime(loan_created_date, "%Y-%m-%d")
                    custom_created_at = custom_created_at.replace(hour=12, minute=0)
            except ValueError:
                flash('Invalid date/time format. Using current date/time.')
                custom_created_at = None
        
        loan = Loan(
            customer_id=customer_id,
            loan_name=loan_name,
            principal_amount=principal,
            remaining_principal=principal,
            interest_rate=interest_rate,
            payment_frequency=payment_frequency,
            loan_type=loan_type,
            admin_notes=admin_notes if admin_notes else None,
            customer_notes=customer_notes if customer_notes else None,
            created_at=custom_created_at if custom_created_at else datetime.utcnow()
        )
        
        db.session.add(loan)
        db.session.commit()
        
        flash('Loan created successfully')
        return redirect(url_for('admin_loans'))
    
    customers = User.query.filter_by(is_admin=False).all()
    return render_template('admin/create_loan.html', customers=customers)

@app.route('/admin/edit-loan/<int:loan_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_loan(loan_id):
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    loan = Loan.query.get_or_404(loan_id)
    
    if request.method == 'POST':
        loan.loan_name = request.form['loan_name']
        new_principal = Decimal(request.form['principal'])
        new_interest_rate = Decimal(request.form['interest_rate']) / 100  # Convert percentage to decimal
        payment_frequency = request.form['payment_frequency']
        loan_type = request.form['loan_type']
        admin_notes = request.form.get('admin_notes', '').strip()  # Get admin notes
        customer_notes = request.form.get('customer_notes', '').strip()  # Get customer notes
        
        # Handle custom creation date
        loan_created_date = request.form.get('loan_created_date', '').strip()
        loan_created_time = request.form.get('loan_created_time', '').strip()
        
        if loan_created_date:
            try:
                from datetime import datetime
                if loan_created_time:
                    # Combine date and time
                    new_created_at = datetime.strptime(f"{loan_created_date} {loan_created_time}", "%Y-%m-%d %H:%M")
                else:
                    # Use date only, default to 12:00 PM
                    new_created_at = datetime.strptime(loan_created_date, "%Y-%m-%d")
                    new_created_at = new_created_at.replace(hour=12, minute=0)
                
                # Update creation date
                loan.created_at = new_created_at
            except ValueError:
                flash('Invalid date/time format. Creation date not changed.')
        
        # Calculate the difference in principal
        principal_difference = new_principal - loan.principal_amount
        
        # Update loan details
        loan.principal_amount = new_principal
        loan.remaining_principal += principal_difference  # Adjust remaining principal
        loan.interest_rate = new_interest_rate
        loan.payment_frequency = payment_frequency
        loan.loan_type = loan_type
        loan.admin_notes = admin_notes if admin_notes else None
        loan.customer_notes = customer_notes if customer_notes else None
        
        # Ensure remaining principal doesn't go negative
        if loan.remaining_principal < 0:
            loan.remaining_principal = Decimal('0')
        
        db.session.commit()
        
        flash('Loan updated successfully')
        return redirect(url_for('admin_loans'))
    
    customers = User.query.filter_by(is_admin=False).all()
    
    # Get all payments for this loan, ordered by date (newest first)
    payments = Payment.query.filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
    
    return render_template('admin/edit_loan.html', 
                         loan=loan, 
                         customers=customers,
                         payments=payments)

@app.route('/admin/loan/<int:loan_id>')
@login_required
def admin_view_loan(loan_id):
    """View detailed loan information (read-only)"""
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    loan = Loan.query.get_or_404(loan_id)
    
    # Get payment history
    payments = Payment.query.filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
    
    # Calculate accumulated interest from loan creation to now
    accumulated_interest = calculate_accumulated_interest(loan)
    
    # Calculate current daily/monthly interest rates (for display purposes)
    daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
    monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
    
    # Calculate days active
    days_active = (date.today() - loan.created_at.date()).days
    
    # Calculate total interest paid for this loan (verified only)
    total_interest_paid = db.session.query(db.func.sum(Payment.interest_amount)).filter_by(loan_id=loan_id, status='verified').scalar() or 0
    
    # Calculate pending payment amounts
    pending_payments = Payment.query.filter_by(loan_id=loan_id, status='pending').all()
    pending_principal = sum(payment.principal_amount for payment in pending_payments)
    pending_interest = sum(payment.interest_amount for payment in pending_payments)
    pending_total = sum(payment.amount for payment in pending_payments)
    
    # Calculate verified payment amounts
    verified_payments = Payment.query.filter_by(loan_id=loan_id, status='verified').all()
    verified_principal = sum(payment.principal_amount for payment in verified_payments)
    verified_interest = sum(payment.interest_amount for payment in verified_payments)
    
    return render_template('admin/view_loan.html', 
                         loan=loan,
                         payments=payments,
                         daily_interest=daily_interest,
                         monthly_interest=monthly_interest,
                         accumulated_interest=accumulated_interest,
                         days_active=days_active,
                         total_interest_paid=total_interest_paid,
                         pending_principal=pending_principal,
                         pending_interest=pending_interest,
                         pending_total=pending_total,
                         verified_principal=verified_principal,
                         verified_interest=verified_interest)

@app.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    users = User.query.order_by(User.created_at.desc()).all()
    total_loans = sum(len(user.loans) for user in users)
    admin_count = sum(1 for user in users if user.is_admin)
    customer_count = len(users) - admin_count
    
    return render_template('admin/users.html', 
                         users=users, 
                         total_loans=total_loans,
                         admin_count=admin_count,
                         customer_count=customer_count)

@app.route('/admin/create-user', methods=['GET', 'POST'])
@login_required
def admin_create_user():
    if not current_user.is_admin:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form.get('email', '').strip()  # Get email, default to empty string
        password = request.form['password']
        is_admin = 'is_admin' in request.form
        
        # Check if username already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return render_template('admin/create_user.html')
        
        # Only check email uniqueness if email is provided
        if email and User.query.filter_by(email=email).first():
            flash('Email already exists')
            return render_template('admin/create_user.html')
        
        # Create new user
        user = User(
            username=username,
            email=email if email else None,  # Store as None if empty
            password_hash=generate_password_hash(password),
            is_admin=is_admin
        )
        
        db.session.add(user)
        db.session.commit()
        
        user_type = 'Admin' if is_admin else 'Customer'
        flash(f'{user_type} user created successfully')
        return redirect(url_for('admin_users'))
    
    return render_template('admin/create_user.html')

# Customer Routes
@app.route('/customer')
@login_required
def customer_dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    
    # Get customer's active loans
    loans = Loan.query.filter_by(customer_id=current_user.id, is_active=True).all()
    
    # Calculate daily interest for each loan
    loan_data = []
    for loan in loans:
        daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
        monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
        
        # Calculate accumulated interest from loan creation to now
        accumulated_interest = calculate_accumulated_interest(loan)
        
        # Calculate pending payment amounts
        pending_payments = Payment.query.filter_by(loan_id=loan.id, status='pending').all()
        pending_principal = sum(payment.principal_amount for payment in pending_payments)
        pending_interest = sum(payment.interest_amount for payment in pending_payments)
        pending_total = sum(payment.amount for payment in pending_payments)
        
        # Calculate verified payment amounts
        verified_payments = Payment.query.filter_by(loan_id=loan.id, status='verified').all()
        verified_principal = sum(payment.principal_amount for payment in verified_payments)
        verified_interest = sum(payment.interest_amount for payment in verified_payments)
        
        loan_data.append({
            'loan': loan,
            'daily_interest': daily_interest,
            'monthly_interest': monthly_interest,
            'accumulated_interest': accumulated_interest,
            'pending_principal': pending_principal,
            'pending_interest': pending_interest,
            'pending_total': pending_total,
            'verified_principal': verified_principal,
            'verified_interest': verified_interest
        })
    
    return render_template('customer/dashboard.html', loan_data=loan_data)

@app.route('/customer/loan/<int:loan_id>')
@login_required
def customer_loan_detail(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.customer_id != current_user.id:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    # Get payment history
    payments = Payment.query.filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
    
    # Calculate accumulated interest from loan creation to now
    accumulated_interest = calculate_accumulated_interest(loan)
    
    # Calculate current daily/monthly interest rates (for display purposes)
    daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
    monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
    
    # Calculate days active
    days_active = (date.today() - loan.created_at.date()).days
    
    # Calculate total interest paid for this loan (verified only)
    total_interest_paid = db.session.query(db.func.sum(Payment.interest_amount)).filter_by(loan_id=loan_id, status='verified').scalar() or 0
    
    # Calculate pending payment amounts
    pending_payments = Payment.query.filter_by(loan_id=loan_id, status='pending').all()
    pending_principal = sum(payment.principal_amount for payment in pending_payments)
    pending_interest = sum(payment.interest_amount for payment in pending_payments)
    pending_total = sum(payment.amount for payment in pending_payments)
    
    # Calculate verified payment amounts
    verified_payments = Payment.query.filter_by(loan_id=loan_id, status='verified').all()
    verified_principal = sum(payment.principal_amount for payment in verified_payments)
    
    return render_template('customer/loan_detail.html', 
                         loan=loan,
                         payments=payments,
                         daily_interest=daily_interest,
                         monthly_interest=monthly_interest,
                         accumulated_interest=accumulated_interest,
                         days_active=days_active,
                         total_interest_paid=total_interest_paid,
                         pending_principal=pending_principal,
                         pending_interest=pending_interest,
                         pending_total=pending_total,
                         verified_principal=verified_principal)

@app.route('/customer/payment/<int:loan_id>', methods=['POST'])
@login_required
def customer_make_payment(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.customer_id != current_user.id:
        flash('Access denied')
        return redirect(url_for('customer_dashboard'))
    
    try:
        payment_amount = Decimal(request.form['amount'])
        transaction_id = request.form.get('transaction_id', '').strip()
        payment_method = request.form.get('payment_method', '')
        payment_date_str = request.form.get('payment_date', '').strip()
        
        if payment_amount <= 0:
            flash('Payment amount must be positive')
            return redirect(url_for('customer_loan_detail', loan_id=loan_id))
        
        if not payment_method:
            flash('Please select a payment method')
            return redirect(url_for('customer_loan_detail', loan_id=loan_id))
        
        # Handle payment date
        payment_date = None
        if payment_date_str:
            try:
                from datetime import datetime
                payment_date = datetime.fromisoformat(payment_date_str.replace('Z', '+00:00'))
            except ValueError:
                flash('Invalid payment date format')
                return redirect(url_for('customer_loan_detail', loan_id=loan_id))
        
        # Handle file upload
        proof_filename = None
        if 'proof_file' in request.files:
            file = request.files['proof_file']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                unique_filename = generate_unique_filename(filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                proof_filename = unique_filename
            elif file and file.filename != '':
                flash('Invalid file type. Please upload JPEG, PNG, GIF, or PDF files only.')
                return redirect(url_for('customer_loan_detail', loan_id=loan_id))
        
        process_payment(loan, payment_amount, payment_date=payment_date, 
                       transaction_id=transaction_id, payment_method=payment_method, 
                       proof_filename=proof_filename)
        
        flash('Payment submitted successfully! Your payment is pending admin verification.')
        
    except ValueError as e:
        flash(str(e))
    except Exception as e:
        flash('Error processing payment')
    
    return redirect(url_for('customer_loan_detail', loan_id=loan_id))

@app.route('/customer/loan/<int:loan_id>/edit-notes', methods=['POST'])
@login_required
def customer_edit_notes(loan_id):
    """Allow customers to edit their loan notes"""
    if current_user.is_admin:
        flash('Access denied. Admin users cannot edit customer notes from this page.')
        return redirect(url_for('admin_dashboard'))
    
    loan = Loan.query.get_or_404(loan_id)
    
    # Ensure the customer can only edit their own loan notes
    if loan.customer_id != current_user.id:
        flash('Access denied. You can only edit notes for your own loans.')
        return redirect(url_for('customer_dashboard'))
    
    if request.method == 'POST':
        customer_notes = request.form.get('customer_notes', '').strip()
        
        loan.customer_notes = customer_notes if customer_notes else None
        db.session.commit()
        
        flash('Notes updated successfully!')
        return redirect(url_for('customer_loan_detail', loan_id=loan_id))
    
    return redirect(url_for('customer_loan_detail', loan_id=loan_id))

# File serving route
@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    """Serve uploaded payment proof files"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# API Routes
@app.route('/api/loan/<int:loan_id>/interest')
@login_required
def api_loan_interest(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.customer_id != current_user.id and not current_user.is_admin:
        return jsonify({'error': 'Access denied'}), 403
    
    daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
    monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
    
    return jsonify({
        'daily_interest': float(daily_interest),
        'monthly_interest': float(monthly_interest),
        'remaining_principal': float(loan.remaining_principal)
    })

@app.route('/api/loan/<int:loan_id>/details')
@login_required
def get_loan_details(loan_id):
    """Get loan details for payment calculation"""
    if not current_user.is_admin:
        return jsonify({'error': 'Access denied'}), 403
    
    loan = Loan.query.get_or_404(loan_id)
    
    # Calculate current interest
    daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
    monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
    
    return jsonify({
        'remaining_principal': float(loan.remaining_principal),
        'interest_rate': float(loan.interest_rate),
        'daily_interest': float(daily_interest),
        'monthly_interest': float(monthly_interest),
        'payment_frequency': loan.payment_frequency
    })

@app.route('/admin/payments')
@login_required
def admin_payments():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.')
        return redirect(url_for('customer_dashboard'))
    
    # Get filter parameters
    customer_filter = request.args.get('customer', '')
    loan_filter = request.args.get('loan', '')
    export = request.args.get('export', '')
    
    # Build query with filters
    query = db.session.query(Payment, Loan, User).join(Loan, Payment.loan_id == Loan.id).join(User, Loan.customer_id == User.id)
    
    if customer_filter:
        query = query.filter(User.username.ilike(f'%{customer_filter}%'))
    if loan_filter:
        query = query.filter(Loan.loan_name.ilike(f'%{loan_filter}%'))
    
    payments = query.order_by(Payment.payment_date.desc()).all()
    
    # Calculate total interest paid
    total_interest_paid = db.session.query(db.func.sum(Payment.interest_amount)).scalar() or 0
    
    # Calculate filtered totals
    filtered_principal = sum(payment.principal_amount for payment, loan, user in payments)
    filtered_interest = sum(payment.interest_amount for payment, loan, user in payments)
    filtered_total = sum(payment.amount for payment, loan, user in payments)
    
    # Handle Excel export
    if export == 'excel':
        return export_payments_to_excel(payments)
    
    # Get unique customers and loans for filter dropdowns
    customers = db.session.query(User.username).filter_by(is_admin=False).distinct().all()
    loans = db.session.query(Loan.loan_name).distinct().all()
    
    return render_template('admin/payments.html', 
                         payments=payments, 
                         total_interest_paid=total_interest_paid,
                         customers=[c[0] for c in customers],
                         loans=[l[0] for l in loans],
                         customer_filter=customer_filter,
                         loan_filter=loan_filter,
                         filtered_principal=filtered_principal,
                         filtered_interest=filtered_interest,
                         filtered_total=filtered_total)

@app.route('/admin/payments/edit/<int:payment_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_payment(payment_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.')
        return redirect(url_for('customer_dashboard'))
    
    payment = Payment.query.get_or_404(payment_id)
    loan = Loan.query.get(payment.loan_id)
    
    if request.method == 'POST':
        # Get form data
        new_amount = request.form.get('amount')
        new_interest_amount = request.form.get('interest_amount')
        new_principal_amount = request.form.get('principal_amount')
        new_transaction_id = request.form.get('transaction_id')
        new_payment_method = request.form.get('payment_method')
        new_status = request.form.get('status')
        
        try:
            # Validate amounts
            new_amount = Decimal(new_amount)
            new_interest_amount = Decimal(new_interest_amount)
            new_principal_amount = Decimal(new_principal_amount)
            
            if new_amount != new_interest_amount + new_principal_amount:
                flash('Total amount must equal interest + principal amount')
                return redirect(url_for('admin_edit_payment', payment_id=payment_id))
            
            # Store old values for loan recalculation
            old_status = payment.status
            old_principal_amount = payment.principal_amount
            
            # Update payment
            payment.amount = new_amount
            payment.interest_amount = new_interest_amount
            payment.principal_amount = new_principal_amount
            payment.transaction_id = new_transaction_id
            payment.payment_method = new_payment_method
            payment.status = new_status
            
            # Handle status changes and balance adjustments
            if new_status == 'verified' and old_status == 'pending':
                # Payment is being verified - reduce balance
                loan.remaining_principal -= new_principal_amount
            elif new_status == 'pending' and old_status == 'verified':
                # Payment is being unverified - add back to balance
                loan.remaining_principal += old_principal_amount
            elif new_status == 'verified' and old_status == 'verified':
                # Payment was already verified, adjust for amount changes
                principal_difference = new_principal_amount - old_principal_amount
                loan.remaining_principal -= principal_difference
            
            # Ensure remaining principal doesn't go negative
            loan.remaining_principal = max(Decimal('0'), loan.remaining_principal)
            
            db.session.commit()
            flash('Payment updated successfully!')
            return redirect(url_for('admin_payments'))
            
        except (ValueError, TypeError) as e:
            flash('Invalid amount format')
            return redirect(url_for('admin_edit_payment', payment_id=payment_id))
    
    return render_template('admin/edit_payment.html', payment=payment, loan=loan)

@app.route('/admin/payments/add', methods=['GET', 'POST'])
@app.route('/admin/payments/add/<int:loan_id>', methods=['GET', 'POST'])
@login_required
def admin_add_payment(loan_id=None):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.')
        return redirect(url_for('customer_dashboard'))
    
    if request.method == 'POST':
        loan_id = request.form.get('loan_id')
        amount = request.form.get('amount')
        transaction_id = request.form.get('transaction_id')
        payment_method = request.form.get('payment_method')
        payment_date_str = request.form.get('payment_date')
        
        try:
            loan = Loan.query.get_or_404(loan_id)
            amount = Decimal(amount)
            
            # Handle payment date
            payment_date = None
            if payment_date_str:
                from datetime import datetime
                payment_date = datetime.fromisoformat(payment_date_str.replace('Z', '+00:00'))
            
            # Process payment (same logic as customer payments)
            process_payment(loan, amount, payment_date=payment_date, 
                          transaction_id=transaction_id, payment_method=payment_method, 
                          proof_filename=None)
            
            flash('Payment added successfully!')
            return redirect(url_for('admin_payments'))
            
        except (ValueError, TypeError) as e:
            flash('Invalid amount format')
            return redirect(url_for('admin_add_payment'))
        except Exception as e:
            flash(f'Error adding payment: {str(e)}')
            return redirect(url_for('admin_add_payment'))
    
    # Get all loans for the dropdown
    loans = Loan.query.all()
    return render_template('admin/add_payment.html', loans=loans, loan_id=loan_id)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create default admin user if it doesn't exist
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@lendingapp.com',
                password_hash=generate_password_hash('admin123'),
                is_admin=True
            )
            db.session.add(admin)
            
            # Create default interest rate
            default_rate = InterestRate(rate=Decimal('0.21'))  # 21%
            db.session.add(default_rate)
            
            db.session.commit()
            print("Default admin user created: username='admin', password='admin123'")
    
    app.run(debug=True)
