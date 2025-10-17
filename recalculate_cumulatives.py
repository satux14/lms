"""
Recalculate Cumulative for All Existing Tracker Entries
========================================================

This script goes through all tracker Excel files and recalculates
cumulative values for entries that have daily_payments but missing cumulative.

Usage:
    python3 recalculate_cumulatives.py
"""

import sys
import os
from pathlib import Path

# Add the app directory to the path
sys.path.insert(0, os.path.dirname(__file__))

from app_multi import app, init_app, VALID_INSTANCES, get_database_uri
import openpyxl

# Import from daily-trackers directory
import importlib.util
spec = importlib.util.spec_from_file_location("tracker_manager", "daily-trackers/tracker_manager.py")
tracker_manager = importlib.util.module_from_spec(spec)
spec.loader.exec_module(tracker_manager)

SHEET_CONFIGS = tracker_manager.SHEET_CONFIGS
TRACKER_TYPES = tracker_manager.TRACKER_TYPES
get_tracker_directory = tracker_manager.get_tracker_directory

def recalculate_tracker_cumulative(instance, filename):
    """Recalculate cumulative for all entries in a tracker"""
    tracker_dir = Path(get_tracker_directory(instance))
    tracker_path = tracker_dir / filename
    
    if not tracker_path.exists():
        return False, f"File not found: {filename}"
    
    try:
        wb = openpyxl.load_workbook(str(tracker_path))
        ws = wb.active
        
        sheet_name = ws.title
        
        # Determine tracker type
        tracker_type = None
        for key, value in TRACKER_TYPES.items():
            if value == sheet_name:
                tracker_type = key
                break
        
        if not tracker_type:
            return False, f"Unknown sheet type: {sheet_name}"
        
        config = SHEET_CONFIGS[tracker_type]
        columns = config['columns']
        data_start = config['data_start_row']
        
        # Get start date for date calculations
        from datetime import timedelta
        start_date_cell = config['start_date_cell']
        start_date = ws[start_date_cell].value
        if isinstance(start_date, str):
            from datetime import datetime
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        changes_made = 0
        cumulative_sum = 0
        
        # Go through all rows and recalculate cumulative and dates
        for row_num in range(data_start, data_start + 400):
            day_cell = f"{columns['day']}{row_num}"
            day_value = ws[day_cell].value
            
            # Stop if we hit empty days
            if day_value is None or day_value == '':
                break
            
            # ALWAYS calculate and set date (force overwrite formulas)
            if start_date and isinstance(day_value, (int, float)):
                calculated_date = start_date + timedelta(days=int(day_value))
                date_cell = f"{columns['date']}{row_num}"
                ws[date_cell] = calculated_date  # Always set
                print(f"  Day {day_value}: Set date to {calculated_date}")
                changes_made += 1
            
            # Get daily payment
            payment_cell = f"{columns['daily_payments']}{row_num}"
            payment_value = ws[payment_cell].value
            
            # Add to cumulative if there's a payment entry (even if 0)
            if payment_value is not None and payment_value != '':
                try:
                    cumulative_sum += float(payment_value)
                except:
                    pass
                
                # ALWAYS set cumulative (force overwrite formulas)
                cumulative_cell = f"{columns['cumulative']}{row_num}"
                ws[cumulative_cell] = cumulative_sum
                print(f"  Day {day_value}: Set cumulative to {cumulative_sum}")
                changes_made += 1
        
        if changes_made > 0:
            wb.save(str(tracker_path))
            return True, f"Updated {changes_made} entries"
        else:
            return True, "No updates needed"
            
    except Exception as e:
        return False, f"Error: {str(e)}"

def main():
    """Main function"""
    print("\n" + "="*70)
    print("RECALCULATE CUMULATIVE VALUES")
    print("="*70)
    
    # Initialize app
    print("\n→ Initializing application...")
    init_app()
    print("✓ Application initialized")
    
    from app_multi import DailyTracker, db
    
    total_updated = 0
    total_trackers = 0
    
    # Run within application context
    with app.app_context():
        for instance in VALID_INSTANCES:
            print(f"\n{'='*70}")
            print(f"Instance: {instance}")
            print(f"{'='*70}")
            
            # Get database URI for this instance
            db_uri = get_database_uri(instance)
            
            # Query trackers directly with the instance database
            from sqlalchemy import create_engine
            from sqlalchemy.orm import sessionmaker
            
            engine = create_engine(db_uri)
            Session = sessionmaker(bind=engine)
            session = Session()
            
            try:
                # Get all active trackers for this instance
                trackers = session.query(DailyTracker).filter_by(is_active=True).all()
                
                if not trackers:
                    print(f"  No trackers found")
                    continue
                
                print(f"  Found {len(trackers)} trackers")
                
                for tracker in trackers:
                    # Get username
                    from app_multi import User
                    user = session.query(User).filter_by(id=tracker.user_id).first()
                    username = user.username if user else "Unknown"
                    
                    print(f"\n  → {tracker.tracker_name} (User: {username})")
                    success, message = recalculate_tracker_cumulative(instance, tracker.filename)
                    
                    if success:
                        print(f"    ✓ {message}")
                        if "Updated" in message:
                            total_updated += 1
                    else:
                        print(f"    ✗ {message}")
                    
                    total_trackers += 1
            finally:
                session.close()
                engine.dispose()
    
    print("\n" + "="*70)
    print("SUMMARY")
    print("="*70)
    print(f"Total trackers checked: {total_trackers}")
    print(f"Trackers updated: {total_updated}")
    print("\n✓ Done! Refresh your browser to see updated cumulative values.")

if __name__ == '__main__':
    main()

