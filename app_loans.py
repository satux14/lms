"""
Loan Routes Module
==================

This module handles all loan-related functionality including:
- Loan management (admin and customer)
- Loan splitting
- Loan cashback configuration
- Loan Excel report generation
"""

from flask import request, redirect, url_for, flash, render_template, send_file, abort, g
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from io import BytesIO

# Import from app_multi - these will be set when register_loan_routes is called
app = None
db = None

# These will be imported from app_multi
VALID_INSTANCES = None
DEFAULT_INSTANCE = None
Loan = None
Payment = None
LoanSplit = None
LoanCashbackConfig = None
User = None
CashbackTransaction = None
get_loan_query = None
get_loan_split_query = None
get_loan_cashback_config_query = None
get_payment_query = None
get_user_query = None
get_cashback_transaction_query = None
add_to_current_instance = None
commit_current_instance = None
get_current_instance_from_g = None
db_manager = None
validate_username_exists = None
get_payment_cashback_total = None

# Helper functions (shared utilities from app_multi)
calculate_daily_interest = None
calculate_monthly_interest = None
calculate_accumulated_interest = None
calculate_interest_for_period = None
DAYS_PER_YEAR = None

# Logging/metrics managers
get_logging_manager = None
get_metrics_manager = None


def register_loan_routes(flask_app, flask_db, valid_instances, default_instance,
                         loan_model, payment_model, loan_split_model, loan_cashback_config_model,
                         user_model, cashback_transaction_model,
                         loan_query_func, loan_split_query_func, loan_cashback_config_query_func,
                         payment_query_func, user_query_func, cashback_transaction_query_func,
                         add_instance_func, commit_instance_func, get_current_instance_func,
                         db_manager_instance, validate_username_exists_helper, get_payment_cashback_total_helper,
                         calculate_daily_interest_func, calculate_monthly_interest_func,
                         calculate_accumulated_interest_func, calculate_interest_for_period_func,
                         days_per_year, get_logging_manager_func, get_metrics_manager_func):
    """Register loan routes with Flask app"""
    global app, db, VALID_INSTANCES, DEFAULT_INSTANCE
    global Loan, Payment, LoanSplit, LoanCashbackConfig, User, CashbackTransaction
    global get_loan_query, get_loan_split_query, get_loan_cashback_config_query
    global get_payment_query, get_user_query, get_cashback_transaction_query
    global add_to_current_instance, commit_current_instance, get_current_instance_from_g
    global db_manager, validate_username_exists, get_payment_cashback_total
    global calculate_daily_interest, calculate_monthly_interest, calculate_accumulated_interest
    global calculate_interest_for_period, DAYS_PER_YEAR
    global get_logging_manager, get_metrics_manager
    
    app = flask_app
    db = flask_db
    VALID_INSTANCES = valid_instances
    DEFAULT_INSTANCE = default_instance
    Loan = loan_model
    Payment = payment_model
    LoanSplit = loan_split_model
    LoanCashbackConfig = loan_cashback_config_model
    User = user_model
    CashbackTransaction = cashback_transaction_model
    get_loan_query = loan_query_func
    get_loan_split_query = loan_split_query_func
    get_loan_cashback_config_query = loan_cashback_config_query_func
    get_payment_query = payment_query_func
    get_user_query = user_query_func
    get_cashback_transaction_query = cashback_transaction_query_func
    add_to_current_instance = add_instance_func
    commit_current_instance = commit_instance_func
    get_current_instance_from_g = get_current_instance_func
    db_manager = db_manager_instance
    validate_username_exists = validate_username_exists_helper
    get_payment_cashback_total = get_payment_cashback_total_helper
    calculate_daily_interest = calculate_daily_interest_func
    calculate_monthly_interest = calculate_monthly_interest_func
    calculate_accumulated_interest = calculate_accumulated_interest_func
    calculate_interest_for_period = calculate_interest_for_period_func
    DAYS_PER_YEAR = days_per_year
    get_logging_manager = get_logging_manager_func
    get_metrics_manager = get_metrics_manager_func
    
    # Register routes
    register_routes()


def get_loan_cashback_total(loan_id, instance_name):
    """Calculate total cashback points given for a loan"""
    try:
        session = db_manager.get_session_for_instance(instance_name)
        total = session.query(
            db.func.sum(CashbackTransaction.points)
        ).filter(
            CashbackTransaction.related_loan_id == loan_id
        ).filter(
            CashbackTransaction.transaction_type.in_(['loan_interest_auto', 'loan_interest_manual'])
        ).scalar() or Decimal('0')
        return total
    except Exception as e:
        print(f"Error calculating loan cashback total: {e}")
        return Decimal('0')


def process_loan_cashback(loan, payment, instance_name, created_by_user_id):
    """Process automatic cashback for a loan payment based on LoanCashbackConfig"""
    try:
        session = db_manager.get_session_for_instance(instance_name)
        
        # Get all active cashback configs for this loan
        configs = session.query(LoanCashbackConfig).filter_by(
            loan_id=loan.id,
            is_active=True
        ).all()
        
        cashback_details = []
        total_cashback = Decimal('0')
        
        for config in configs:
            if config.cashback_type == 'percentage':
                # Calculate points as percentage of interest amount
                points = payment.interest_amount * config.cashback_value
            else:  # fixed
                # Use fixed amount
                points = config.cashback_value
            
            # Round to 2 decimal places
            points = points.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            if points > 0:
                # Create cashback transaction
                transaction = CashbackTransaction(
                    from_user_id=None,  # System/admin grant
                    to_user_id=config.user_id,
                    points=points,
                    transaction_type='loan_interest_auto',
                    related_loan_id=loan.id,
                    related_payment_id=payment.id,
                    notes=f"Automatic cashback from loan '{loan.loan_name}' payment",
                    created_by_user_id=created_by_user_id
                )
                session.add(transaction)
                total_cashback += points
                cashback_details.append({
                    'user': config.user.username,
                    'points': float(points),
                    'type': config.cashback_type
                })
        
        session.commit()
        
        # Log cashback activity if any cashback was given
        if total_cashback > 0:
            try:
                logging_mgr = get_logging_manager(instance_name)
                # Get username from created_by_user_id
                created_by_user = session.query(User).filter_by(id=created_by_user_id).first()
                username = created_by_user.username if created_by_user else 'system'
                
                logging_mgr.log_activity(
                    action='cashback_loan_auto',
                    username=username,
                    user_id=created_by_user_id,
                    resource_type='loan',
                    resource_id=loan.id,
                    details={
                        'loan_name': loan.loan_name,
                        'payment_id': payment.id,
                        'total_cashback': float(total_cashback),
                        'cashback_details': cashback_details
                    }
                )
            except Exception as log_error:
                print(f"[ERROR] Failed to log loan cashback: {log_error}")
        
    except Exception as e:
        print(f"Error processing loan cashback: {e}")
        session.rollback()


def generate_loan_calculation_excel(loan):
    """Generate Excel report showing daily loan calculations for 6 months"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Sheet 1: With Payments - Database (Static from DB)
    ws1 = wb.active
    ws1.title = "DB Payments (Actual)"
    
    # Sheet 2: With Payments - Manual Entry (Formulas)
    ws2 = wb.create_sheet("Manual Payments")
    
    # Sheet 3: Without Payments (Projection)
    ws3 = wb.create_sheet("Without Payments")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all payments for this loan (sorted by date)
    payments = get_payment_query().filter_by(loan_id=loan.id, status='verified').order_by(Payment.payment_date).all()
    
    # Calculate for 1 year (365 days)
    num_days = 365
    start_date = loan.created_at.date()
    
    def setup_sheet(ws, include_payments=True):
        """Setup headers and styling for a sheet"""
        headers = [
            "Day", "Date", "Opening Principal", "Daily Interest", 
            "Accumulated Interest", "Payment Amount", "Interest Paid", 
            "Principal Paid", "Closing Principal", "Remaining Interest"
        ]
        
        if not include_payments:
            headers = [h for h in headers if h not in ["Payment Amount", "Interest Paid", "Principal Paid"]]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        return headers
    
    # Setup all three sheets
    headers_db = setup_sheet(ws1, include_payments=True)
    headers_manual = setup_sheet(ws2, include_payments=True)
    headers_without = setup_sheet(ws3, include_payments=False)
    
    # Get all payments for sheets with database data
    # Handle interest rate stored as decimal (0.12) or percentage (12)
    interest_rate_value = float(loan.interest_rate)
    if interest_rate_value < 1:
        # Stored as decimal (0.12 for 12%), don't divide by 100
        daily_rate = interest_rate_value / DAYS_PER_YEAR
    else:
        # Stored as percentage (12 for 12%), divide by 100
        daily_rate = interest_rate_value / DAYS_PER_YEAR / 100
    
    # Pre-populate payment amounts from database
    payment_dict = {}
    for payment in payments:
        payment_date_key = payment.payment_date.date()
        if payment_date_key not in payment_dict:
            payment_dict[payment_date_key] = {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')}
        payment_dict[payment_date_key]['amount'] += payment.amount
        payment_dict[payment_date_key]['interest'] += payment.interest_amount
        payment_dict[payment_date_key]['principal'] += payment.principal_amount
    
    # ===== SHEET 1: Database Payments (Static from DB) =====
    opening_principal = loan.principal_amount
    accumulated_interest = Decimal('0')
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2
        
        # Daily interest on opening principal
        daily_interest = opening_principal * Decimal(str(daily_rate))
        accumulated_interest += daily_interest
        
        # Get payments from database for this date
        payment_data = payment_dict.get(current_date, {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')})
        payment_amount = payment_data['amount']
        interest_paid = payment_data['interest']
        principal_paid = payment_data['principal']
        
        # Update accumulated interest and principal
        accumulated_interest -= interest_paid
        closing_principal = opening_principal - principal_paid
        
        # Write data to sheet (static values from database)
        ws1.cell(row=row, column=1, value=day + 1)
        ws1.cell(row=row, column=2, value=current_date.strftime('%Y-%m-%d'))
        ws1.cell(row=row, column=3, value=float(opening_principal)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=4, value=float(daily_interest)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=5, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=6, value=float(payment_amount)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=7, value=float(interest_paid)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=8, value=float(principal_paid)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=9, value=float(closing_principal)).number_format = '₹#,##0.00'
        ws1.cell(row=row, column=10, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        
        # Apply borders
        for col in range(1, 11):
            ws1.cell(row=row, column=col).border = border
        
        # Highlight payment rows
        if payment_amount > 0:
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Update for next day
        opening_principal = closing_principal
    
    # ===== SHEET 2: Manual Payments (Formulas for manual entry) =====
    # This sheet has formulas so you can manually enter payments and see calculations update
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2  # Starting at row 2 for data (headers will be inserted later, shifting to row 7)
        final_row = row + 5  # Account for 5 header rows that will be inserted
        
        # Get pre-populated payment data
        payment_data = payment_dict.get(current_date, {'amount': Decimal('0'), 'interest': Decimal('0'), 'principal': Decimal('0')})
        
        # Write static values
        ws2.cell(row=row, column=1, value=day + 1)  # Day
        ws2.cell(row=row, column=2, value=current_date.strftime('%Y-%m-%d'))  # Date
        
        # Opening Principal (formula: previous day's closing or initial value)
        if day == 0:
            # First day: use initial principal as value
            ws2.cell(row=row, column=3, value=float(loan.principal_amount)).number_format = '₹#,##0.00'
        else:
            # Subsequent days: reference previous day's closing with error handling
            ws2.cell(row=row, column=3, value=f'=IFERROR(I{final_row-1},{float(loan.principal_amount)})').number_format = '₹#,##0.00'
        
        # Daily Interest (formula: Opening Principal * daily rate)
        ws2.cell(row=row, column=4, value=f'=IFERROR(C{final_row}*{daily_rate},0)').number_format = '₹#,##0.00'
        
        # Payment Amount (editable - pre-populate from database) - MOVED BEFORE accumulated interest
        ws2.cell(row=row, column=6, value=float(payment_data['amount'])).number_format = '₹#,##0.00'
        
        # Accumulated Interest BEFORE payment (formula: previous remaining + today's interest)
        if day == 0:
            # First day: just today's interest
            ws2.cell(row=row, column=5, value=f'=IFERROR(D{final_row},0)').number_format = '₹#,##0.00'
        else:
            # Subsequent days: previous day's REMAINING interest (J) + today's new interest
            ws2.cell(row=row, column=5, value=f'=IFERROR(J{final_row-1}+D{final_row},0)').number_format = '₹#,##0.00'
        
        # Interest Paid (formula: MIN(payment amount, accumulated interest before payment))
        ws2.cell(row=row, column=7, value=f'=IFERROR(MIN(F{final_row},E{final_row}),0)').number_format = '₹#,##0.00'
        
        # Principal Paid (formula: payment amount - interest paid)
        ws2.cell(row=row, column=8, value=f'=IFERROR(F{final_row}-G{final_row},0)').number_format = '₹#,##0.00'
        
        # Closing Principal (formula: opening principal - principal paid)
        ws2.cell(row=row, column=9, value=f'=IFERROR(C{final_row}-H{final_row},0)').number_format = '₹#,##0.00'
        
        # Remaining Interest (formula: accumulated interest AFTER payment = accumulated - interest paid)
        ws2.cell(row=row, column=10, value=f'=IFERROR(E{final_row}-G{final_row},0)').number_format = '₹#,##0.00'
        
        # Apply borders
        for col in range(1, 11):
            ws2.cell(row=row, column=col).border = border
        
        # Highlight payment rows (from database - can be cleared for manual entry)
        if payment_data['amount'] > 0:
            for col in range(1, 11):
                ws2.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # ===== SHEET 3: Without Payments (Projection) =====
    # Static projection showing interest accumulation without any payments
    opening_principal = loan.principal_amount
    accumulated_interest = Decimal('0')
    
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        row = day + 2
        
        # Daily interest on opening principal (no reduction since no payments)
        daily_interest = opening_principal * Decimal(str(daily_rate))
        accumulated_interest += daily_interest
        
        # Write data to sheet
        col = 1
        ws3.cell(row=row, column=col, value=day + 1)
        col += 1
        ws3.cell(row=row, column=col, value=current_date.strftime('%Y-%m-%d'))
        col += 1
        ws3.cell(row=row, column=col, value=float(opening_principal)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(daily_interest)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(opening_principal)).number_format = '₹#,##0.00'
        col += 1
        ws3.cell(row=row, column=col, value=float(accumulated_interest)).number_format = '₹#,##0.00'
        
        # Apply borders
        for c in range(1, len(headers_without) + 1):
            ws3.cell(row=row, column=c).border = border
    
    # Add summary information at the top of all sheets
    for ws in [ws1, ws2, ws3]:
        ws.insert_rows(1, 5)
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Loan Calculation Report: {loan.loan_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Handle interest rate display (stored as decimal or percentage)
        interest_rate_display = float(loan.interest_rate)
        if interest_rate_display < 1:
            interest_rate_display = interest_rate_display * 100
        
        ws['A2'] = f"Customer: {loan.customer.username}"
        ws['A3'] = f"Principal Amount: ₹{float(loan.principal_amount):,.2f}"
        ws['A4'] = f"Interest Rate: {interest_rate_display:.2f}% per annum"
        ws['A5'] = f"Loan Created: {loan.created_at.strftime('%Y-%m-%d')}"
        
        # Move headers down
        for row in ws.iter_rows(min_row=6, max_row=6):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Re-apply column widths after inserting rows (to ensure they're preserved)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def register_routes():
    """Register all loan routes"""
    
    # Admin Loan Routes
    @app.route('/<instance_name>/admin/loans')
    @login_required
    def admin_loans(instance_name):
        """Admin loans page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        # Get filter parameters
        loan_type = request.args.get('loan_type', '')
        frequency = request.args.get('frequency', '')
        customer_id = request.args.get('customer', '')
        loan_name = request.args.get('loan_name', '')
        min_rate = request.args.get('min_rate', '')
        max_rate = request.args.get('max_rate', '')
        status = request.args.get('status', '')
        
        # Build query
        query = get_loan_query()
        
        if loan_type:
            query = query.filter(Loan.loan_type == loan_type)
        if frequency:
            query = query.filter(Loan.payment_frequency == frequency)
        if customer_id:
            query = query.filter(Loan.customer_id == customer_id)
        if loan_name:
            query = query.filter(Loan.loan_name.contains(loan_name))
        if min_rate:
            # Convert percentage to decimal for comparison
            min_rate_decimal = float(min_rate) / 100
            query = query.filter(Loan.interest_rate >= min_rate_decimal)
        if max_rate:
            # Convert percentage to decimal for comparison
            max_rate_decimal = float(max_rate) / 100
            query = query.filter(Loan.interest_rate <= max_rate_decimal)
        if status:
            if status == 'active':
                query = query.filter(Loan.is_active == True, Loan.status == 'active')
            elif status == 'closed':
                query = query.filter(Loan.status == 'closed')
            elif status == 'paid_off':
                query = query.filter(Loan.is_active == False, Loan.status != 'closed')
            elif status == 'all':
                # Show all loans including closed
                pass  # Don't filter by status
        else:
            # Default: show only active loans (exclude closed)
            query = query.filter(Loan.is_active == True, Loan.status == 'active')
        
        # Get sorting parameters
        sort_by = request.args.get('sort', 'created_at')
        sort_order = request.args.get('order', 'desc')
        
        if sort_by == 'customer':
            if sort_order == 'asc':
                query = query.join(User).order_by(User.username.asc())
            else:
                query = query.join(User).order_by(User.username.desc())
        elif sort_by == 'loan_name':
            if sort_order == 'asc':
                query = query.order_by(Loan.loan_name.asc())
            else:
                query = query.order_by(Loan.loan_name.desc())
        elif sort_by == 'principal':
            if sort_order == 'asc':
                query = query.order_by(Loan.principal_amount.asc())
            else:
                query = query.order_by(Loan.principal_amount.desc())
        elif sort_by == 'rate':
            if sort_order == 'asc':
                query = query.order_by(Loan.interest_rate.asc())
            else:
                query = query.order_by(Loan.interest_rate.desc())
        else:  # created_at
            if sort_order == 'asc':
                query = query.order_by(Loan.created_at.asc())
            else:
                query = query.order_by(Loan.created_at.desc())
        
        loans = query.all()
        customers = get_user_query().filter_by(is_admin=False).all()
        
        # Calculate interest paid, interest pending, monthly interest, and cashback for each loan
        loans_with_interest = []
        total_cashback = Decimal('0')
        total_interest_pending = Decimal('0')
        for loan in loans:
            # Get verified interest payments for this loan
            interest_paid = get_payment_query().filter_by(
                loan_id=loan.id, 
                status='verified'
            ).with_entities(db.func.sum(Payment.interest_amount)).scalar() or 0
            
            # Calculate monthly interest
            monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
            
            # Calculate interest pending
            interest_data = calculate_accumulated_interest(loan)
            interest_pending = interest_data['monthly'] if loan.payment_frequency == 'monthly' else interest_data['daily']
            total_interest_pending += interest_pending
            
            # Calculate cashback total for this loan
            loan_cashback = get_loan_cashback_total(loan.id, instance_name)
            total_cashback += loan_cashback
            
            # Get last paid date (most recent verified payment)
            verified_payments = get_payment_query().filter_by(loan_id=loan.id, status='verified').all()
            last_paid_date = None
            last_paid_amount = None
            if verified_payments:
                last_payment = max(verified_payments, key=lambda p: p.payment_date)
                last_paid_date = last_payment.payment_date.date()
                last_paid_amount = last_payment.amount
            
            loans_with_interest.append({
                'loan': loan,
                'interest_paid': interest_paid,
                'monthly_interest': monthly_interest,
                'interest_pending': interest_pending,
                'cashback_total': loan_cashback,
                'last_paid_date': last_paid_date,
                'last_paid_amount': last_paid_amount
            })
        
        return render_template('admin/loans.html', 
                             loans=loans_with_interest, 
                             customers=customers,
                             loan_type=loan_type,
                             frequency=frequency,
                             customer_id=customer_id,
                             loan_name=loan_name,
                             min_rate=min_rate,
                             max_rate=max_rate,
                             status=status,
                             sort_by=sort_by,
                             sort_order=sort_order,
                             total_cashback=float(total_cashback),
                             total_interest_pending=float(total_interest_pending),
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/create-loan', methods=['GET', 'POST'])
    @login_required
    def admin_create_loan(instance_name):
        """Admin create loan page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        if request.method == 'POST':
            customer_id = request.form['customer_id']
            loan_name = request.form['loan_name']
            principal_amount = Decimal(request.form['principal'])
            interest_rate = Decimal(request.form['interest_rate']) / 100
            payment_frequency = request.form['payment_frequency']
            loan_type = request.form['loan_type']
            admin_notes = request.form.get('admin_notes', '')
            customer_notes = request.form.get('customer_notes', '')
            # Handle custom creation date if provided
            custom_created_date = request.form.get('loan_created_date')
            custom_created_time = request.form.get('loan_created_time')
            
            if custom_created_date:
                try:
                    if custom_created_time:
                        # Combine date and time
                        created_at_str = f"{custom_created_date} {custom_created_time}"
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M')
                    else:
                        # Use date only, set time to 00:00
                        created_at = datetime.strptime(custom_created_date, '%Y-%m-%d')
                except ValueError:
                    flash('Invalid date format. Using current date.', 'warning')
                    created_at = datetime.utcnow()
            elif custom_created_time:
                try:
                    # If only time is provided, use current date with specified time
                    # Parse time string manually to avoid datetime.strptime() issue with time-only format
                    time_parts = custom_created_time.split(':')
                    if len(time_parts) == 2:
                        hour = int(time_parts[0])
                        minute = int(time_parts[1])
                        created_at = datetime.utcnow().replace(hour=hour, minute=minute, second=0, microsecond=0)
                    else:
                        raise ValueError("Invalid time format")
                except (ValueError, IndexError):
                    flash('Invalid time format. Using current date/time.', 'warning')
                    created_at = datetime.utcnow()
            else:
                created_at = datetime.utcnow()
            
            loan = Loan(
                customer_id=customer_id,
                loan_name=loan_name,
                principal_amount=principal_amount,
                remaining_principal=principal_amount,
                interest_rate=interest_rate,
                payment_frequency=payment_frequency,
                loan_type=loan_type,
                admin_notes=admin_notes,
                customer_notes=customer_notes,
                created_at=created_at
            )
            add_to_current_instance(loan)
            
            flash('Loan created successfully')
            return redirect(url_for('admin_loans', instance_name=instance_name))
        
        customers = get_user_query().filter_by(is_admin=False).all()
        return render_template('admin/create_loan.html', 
                             customers=customers,
                             current_date=date.today().strftime('%Y-%m-%d'),
                             current_time=datetime.now().strftime('%H:%M'),
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/edit-loan/<int:loan_id>', methods=['GET', 'POST'])
    @login_required
    def admin_edit_loan(instance_name, loan_id):
        """Admin edit loan page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        if request.method == 'POST':
            loan.loan_name = request.form['loan_name']
            loan.principal_amount = Decimal(request.form['principal'])
            # remaining_principal is calculated automatically, not edited directly
            loan.interest_rate = Decimal(request.form['interest_rate']) / 100
            loan.payment_frequency = request.form['payment_frequency']
            loan.loan_type = request.form['loan_type']
            loan.admin_notes = request.form.get('admin_notes', '')
            loan.customer_notes = request.form.get('customer_notes', '')
            
            # Handle custom creation date if provided
            custom_created_date = request.form.get('loan_created_date')
            custom_created_time = request.form.get('loan_created_time')
            
            if custom_created_date:
                try:
                    if custom_created_time:
                        # Combine date and time
                        created_at_str = f"{custom_created_date} {custom_created_time}"
                        loan.created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M')
                    else:
                        # Use date only, set time to 00:00
                        loan.created_at = datetime.strptime(custom_created_date, '%Y-%m-%d')
                except ValueError:
                    flash('Invalid date format. Using current date.', 'warning')
                    # Keep existing created_at if date is invalid
            elif custom_created_time:
                # If only time is provided, update the time part of existing date
                try:
                    # Parse time string manually to avoid datetime.strptime() issue with time-only format
                    time_parts = custom_created_time.split(':')
                    if len(time_parts) == 2:
                        hour = int(time_parts[0])
                        minute = int(time_parts[1])
                        loan.created_at = loan.created_at.replace(hour=hour, minute=minute, second=0, microsecond=0)
                    else:
                        raise ValueError("Invalid time format")
                except (ValueError, IndexError):
                    flash('Invalid time format. Time not updated.', 'warning')
            
            commit_current_instance()
            flash('Loan updated successfully')
            return redirect(url_for('admin_loans', instance_name=instance_name))
        
        # Get payment history for this loan
        payments = get_payment_query().filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
        
        return render_template('admin/edit_loan.html', 
                             loan=loan,
                             payments=payments,
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/loan/<int:loan_id>')
    @login_required
    def admin_view_loan(instance_name, loan_id):
        """Admin view loan details page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        if not loan:
            flash('Loan not found', 'error')
            return redirect(url_for('admin_loans', instance_name=instance_name))
        
        # For closed loans, use 0 for remaining principal in calculations
        if loan.status == 'closed':
            effective_remaining_principal = Decimal('0')
        else:
            effective_remaining_principal = loan.remaining_principal
        
        # Calculate interest information
        daily_interest = calculate_daily_interest(effective_remaining_principal, loan.interest_rate)
        monthly_interest = calculate_monthly_interest(effective_remaining_principal, loan.interest_rate)
        interest_data = calculate_accumulated_interest(loan)
        accumulated_interest_daily = interest_data['daily']
        accumulated_interest_monthly = interest_data['monthly']
        
        # Get payment history
        payments = get_payment_query().filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
        
        # Calculate verified amounts
        verified_payments = [p for p in payments if p.status == 'verified']
        verified_principal = sum(payment.principal_amount for payment in verified_payments)
        verified_interest = sum(payment.interest_amount for payment in verified_payments)
        
        # Calculate pending amounts
        pending_payments = [p for p in payments if p.status == 'pending']
        pending_principal = sum(payment.principal_amount for payment in pending_payments)
        pending_interest = sum(payment.interest_amount for payment in pending_payments)
        pending_total = pending_principal + pending_interest
        
        # Calculate days active
        days_active = (date.today() - loan.created_at.date()).days
        
        # Calculate total cashback given for this loan
        loan_cashback_total = get_loan_cashback_total(loan_id, instance_name)
        
        # Get cashback for each payment
        payment_cashback_map = {}
        payment_cashback_recipients_map = {}
        for payment in payments:
            payment_id = payment.id
            payment_cashback_map[payment_id] = get_payment_cashback_total(payment_id, instance_name)
            
            # Get cashback recipients for this payment
            session = db_manager.get_session_for_instance(instance_name)
            cashback_transactions = session.query(CashbackTransaction).filter_by(
                related_payment_id=payment_id
            ).all()
            recipients = []
            for txn in cashback_transactions:
                if txn.to_user:
                    recipients.append({
                        'username': txn.to_user.username,
                        'points': float(txn.points)
                    })
            payment_cashback_recipients_map[payment_id] = recipients
        
        # Get cashback configs for this loan
        cashback_configs = get_loan_cashback_config_query().filter_by(
            loan_id=loan_id,
            is_active=True
        ).all()
        
        # Get all users for assignment - any user (customer or admin) can be assigned as moderator
        # Exclude the current admin user to avoid self-assignment
        all_moderators = get_user_query().filter(User.id != current_user.id).order_by(User.username).all()
        
        # Get split loans for this loan
        splits = get_loan_split_query().filter_by(original_loan_id=loan_id).all()
        split_loans = []
        total_split_principal = Decimal('0')
        for split in splits:
            split_loan = get_loan_query().filter_by(id=split.split_loan_id).first()
            if split_loan:
                split_loans.append({
                    'split': split,
                    'loan': split_loan
                })
                total_split_principal += split.split_principal_amount
        
        # Calculate principal after split
        principal_after_split = loan.principal_amount - total_split_principal
        
        # Get split loan info for payments
        payment_split_loan_map = {}
        for payment in payments:
            if payment.split_loan_id:
                split_loan = get_loan_query().filter_by(id=payment.split_loan_id).first()
                if split_loan:
                    payment_split_loan_map[payment.id] = split_loan
        
        return render_template('admin/view_loan.html', 
                             loan=loan,
                             daily_interest=daily_interest,
                             monthly_interest=monthly_interest,
                             accumulated_interest_daily=accumulated_interest_daily,
                             accumulated_interest_monthly=accumulated_interest_monthly,
                             interest_data=interest_data,
                             verified_principal=verified_principal,
                             verified_interest=verified_interest,
                             loan_cashback_total=loan_cashback_total,
                             pending_principal=pending_principal,
                             pending_interest=pending_interest,
                             pending_total=pending_total,
                             payments=payments,
                             days_active=days_active,
                             all_moderators=all_moderators,
                             current_date=date.today(),
                             payment_cashback_map=payment_cashback_map,
                             payment_cashback_recipients_map=payment_cashback_recipients_map,
                             cashback_configs=cashback_configs,
                             split_loans=split_loans,
                             payment_split_loan_map=payment_split_loan_map,
                             principal_after_split=principal_after_split,
                             total_split_principal=total_split_principal,
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/loan/<int:loan_id>/excel')
    @login_required
    def admin_loan_excel(instance_name, loan_id):
        """Generate and download Excel report for loan calculations"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        try:
            # Generate Excel file
            output = generate_loan_calculation_excel(loan)
            
            # Create filename
            filename = f"loan_calculation_{loan.loan_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Send file
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            flash(f'Error generating Excel report: {str(e)}')
            return redirect(url_for('admin_view_loan', instance_name=instance_name, loan_id=loan_id))


    @app.route('/<instance_name>/admin/close-loan/<int:loan_id>', methods=['POST'])
    @login_required
    def admin_close_loan(instance_name, loan_id):
        """Admin close loan for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        try:
            # Close the loan
            loan.status = 'closed'
            loan.is_active = False
            commit_current_instance()
            
            flash(f'Loan "{loan.loan_name}" has been closed successfully')
            return redirect(url_for('admin_loans', instance_name=instance_name))
            
        except Exception as e:
            flash(f'Error closing loan: {str(e)}')
            return redirect(url_for('admin_loans', instance_name=instance_name))


    @app.route('/<instance_name>/admin/delete-loan/<int:loan_id>', methods=['POST'])
    @login_required
    def admin_delete_loan(instance_name, loan_id):
        """Admin delete loan for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        try:
            # Check if loan has any payments
            payments_count = get_payment_query().filter_by(loan_id=loan_id).count()
            if payments_count > 0:
                flash('Cannot delete loan with existing payments', 'error')
                return redirect(url_for('admin_loans', instance_name=instance_name))
            
            # Delete the loan
            loan_name = loan.loan_name
            get_loan_query().filter_by(id=loan_id).delete()
            commit_current_instance()
            
            flash(f'Loan "{loan_name}" has been deleted successfully')
            return redirect(url_for('admin_loans', instance_name=instance_name))
            
        except Exception as e:
            flash(f'Error deleting loan: {str(e)}')
            return redirect(url_for('admin_loans', instance_name=instance_name))


    @app.route('/<instance_name>/admin/split-loan/<int:loan_id>', methods=['GET', 'POST'])
    @login_required
    def admin_split_loan(instance_name, loan_id):
        """Admin split loan - create a new loan with part of the principal"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        if request.method == 'POST':
            split_principal = Decimal(request.form['split_principal'])
            split_loan_name = request.form.get('split_loan_name', f'{loan.loan_name} - Split')
            split_interest_rate = Decimal(request.form.get('split_interest_rate', loan.interest_rate * 100)) / 100
            
            # Validate split amount
            if split_principal <= 0:
                flash('Split principal amount must be greater than 0', 'error')
                return redirect(url_for('admin_split_loan', instance_name=instance_name, loan_id=loan_id))
            
            if split_principal >= loan.remaining_principal:
                flash(f'Split amount (₹{split_principal:.2f}) must be less than remaining principal (₹{loan.remaining_principal:.2f})', 'error')
                return redirect(url_for('admin_split_loan', instance_name=instance_name, loan_id=loan_id))
            
            # Validate interest rate
            if split_interest_rate < 0 or split_interest_rate > 1:
                flash('Interest rate must be between 0% and 100%', 'error')
                return redirect(url_for('admin_split_loan', instance_name=instance_name, loan_id=loan_id))
            
            try:
                # Create new loan for the split amount
                split_loan = Loan(
                    customer_id=loan.customer_id,
                    loan_name=split_loan_name,
                    principal_amount=split_principal,
                    remaining_principal=split_principal,  # Initially, no payments assigned yet
                    interest_rate=split_interest_rate,
                    payment_frequency=loan.payment_frequency,
                    loan_type=loan.loan_type,
                    admin_notes=f'Split from loan #{loan.id}: {loan.loan_name}',
                    customer_notes=loan.customer_notes,
                    created_at=datetime.utcnow(),
                    is_active=True,
                    status='active'
                )
                add_to_current_instance(split_loan)
                commit_current_instance()
                
                # Create LoanSplit record
                loan_split = LoanSplit(
                    original_loan_id=loan.id,
                    split_loan_id=split_loan.id,
                    split_principal_amount=split_principal,
                    created_by_user_id=current_user.id
                )
                add_to_current_instance(loan_split)
                commit_current_instance()
                
                # Reduce the original loan's remaining principal
                loan.remaining_principal -= split_principal
                commit_current_instance()
                
                flash(f'Loan split successfully! New loan "{split_loan_name}" (ID: {split_loan.id}) created with ₹{split_principal:.2f} principal at {split_interest_rate * 100:.2f}% interest rate.', 'success')
                return redirect(url_for('admin_view_loan', instance_name=instance_name, loan_id=loan.id))
                
            except Exception as e:
                flash(f'Error splitting loan: {str(e)}', 'error')
                return redirect(url_for('admin_split_loan', instance_name=instance_name, loan_id=loan_id))
        
        # GET request - show split form
        # Get existing splits for this loan
        splits = get_loan_split_query().filter_by(original_loan_id=loan_id).all()
        
        return render_template('admin/split_loan.html',
                             loan=loan,
                             splits=splits,
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/assign-payment-to-split/<int:payment_id>', methods=['GET', 'POST'])
    @login_required
    def admin_assign_payment_to_split(instance_name, payment_id):
        """Admin assign payment to a split loan"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        payment = get_payment_query().filter_by(id=payment_id).first() or abort(404)
        loan = payment.loan
        
        if request.method == 'POST':
            split_loan_id = request.form.get('split_loan_id')
            
            if not split_loan_id:
                flash('Please select a split loan', 'error')
                return redirect(url_for('admin_assign_payment_to_split', instance_name=instance_name, payment_id=payment_id))
            
            split_loan = get_loan_query().filter_by(id=split_loan_id).first()
            if not split_loan:
                flash('Split loan not found', 'error')
                return redirect(url_for('admin_assign_payment_to_split', instance_name=instance_name, payment_id=payment_id))
            
            # Verify this is a split of the original loan
            loan_split = get_loan_split_query().filter_by(
                original_loan_id=loan.id,
                split_loan_id=split_loan_id
            ).first()
            
            if not loan_split:
                flash('Selected loan is not a split of this loan', 'error')
                return redirect(url_for('admin_assign_payment_to_split', instance_name=instance_name, payment_id=payment_id))
            
            try:
                # Store original principal amount at time of payment
                payment.original_principal_amount = payment.principal_amount
                # Assign payment to split loan
                payment.split_loan_id = split_loan_id
                
                # Update split loan's remaining principal
                if payment.status == 'verified':
                    split_loan.remaining_principal -= payment.principal_amount
                    if split_loan.remaining_principal < 0:
                        split_loan.remaining_principal = Decimal('0')
                    
                    # If split loan is fully paid, close it
                    if split_loan.remaining_principal <= Decimal('0.01'):
                        split_loan.status = 'closed'
                        split_loan.is_active = False
                
                commit_current_instance()
                
                flash(f'Payment assigned to split loan "{split_loan.loan_name}" successfully!', 'success')
                return redirect(url_for('admin_view_loan', instance_name=instance_name, loan_id=loan.id))
                
            except Exception as e:
                flash(f'Error assigning payment: {str(e)}', 'error')
                return redirect(url_for('admin_assign_payment_to_split', instance_name=instance_name, payment_id=payment_id))
        
        # GET request - show assignment form
        # Get all split loans for this loan
        splits = get_loan_split_query().filter_by(original_loan_id=loan.id).all()
        split_loans = [get_loan_query().filter_by(id=split.split_loan_id).first() for split in splits]
        split_loans = [sl for sl in split_loans if sl and sl.is_active]
        
        return render_template('admin/assign_payment_to_split.html',
                             payment=payment,
                             loan=loan,
                             split_loans=split_loans,
                             instance_name=instance_name)


    @app.route('/<instance_name>/admin/loan/<int:loan_id>/assign-moderator/<int:moderator_id>', methods=['POST'])
    @login_required
    def admin_assign_moderator_to_loan(instance_name, loan_id, moderator_id):
        """Assign a moderator to a loan"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied', 'error')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first()
        # Allow any user (customer or admin) to be assigned as moderator - no need for is_moderator flag
        moderator = get_user_query().filter_by(id=moderator_id).first()
        
        if not loan or not moderator:
            flash('Loan or user not found', 'error')
            return redirect(url_for('admin_loans', instance_name=instance_name))
        
        if moderator not in loan.assigned_moderators:
            loan.assigned_moderators.append(moderator)
            commit_current_instance()
            flash(f'{moderator.username} assigned to loan: {loan.loan_name}', 'success')
        else:
            flash(f'{moderator.username} is already assigned to this loan', 'info')
        
        return redirect(url_for('admin_view_loan', instance_name=instance_name, loan_id=loan_id))


    @app.route('/<instance_name>/admin/loan/<int:loan_id>/unassign-moderator/<int:moderator_id>', methods=['POST'])
    @login_required
    def admin_unassign_moderator_from_loan(instance_name, loan_id, moderator_id):
        """Unassign a moderator from a loan"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if not current_user.is_admin:
            flash('Access denied', 'error')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first()
        moderator = get_user_query().filter_by(id=moderator_id).first()
        
        if not loan or not moderator:
            flash('Loan or moderator not found', 'error')
            return redirect(url_for('admin_loans', instance_name=instance_name))
        
        if moderator in loan.assigned_moderators:
            loan.assigned_moderators.remove(moderator)
            commit_current_instance()
            flash(f'{moderator.username} unassigned from loan: {loan.loan_name}', 'success')
        else:
            flash(f'{moderator.username} is not assigned to this loan', 'info')
        
        return redirect(url_for('admin_view_loan', instance_name=instance_name, loan_id=loan_id))

    # Customer Loan Routes
    @app.route('/<instance_name>/customer/loans')
    @login_required
    def customer_all_loans(instance_name):
        """Customer view all loans page"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard', instance_name=instance_name))
        
        # Get customer's active loans only
        loans = get_loan_query().filter_by(customer_id=current_user.id, is_active=True).all()
        
        loan_data = []
        for loan in loans:
            daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
            monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
            interest_data = calculate_accumulated_interest(loan)
            accumulated_interest_daily = interest_data['daily']
            accumulated_interest_monthly = interest_data['monthly']
            
            # Calculate pending payments for this specific loan
            pending_payments = get_payment_query().filter_by(loan_id=loan.id, status='pending').all()
            pending_principal = sum(payment.principal_amount for payment in pending_payments)
            pending_interest = sum(payment.interest_amount for payment in pending_payments)
            pending_total = sum(payment.amount for payment in pending_payments)
            
            # Calculate verified payments for this specific loan
            verified_payments = get_payment_query().filter_by(loan_id=loan.id, status='verified').all()
            verified_principal = sum(payment.principal_amount for payment in verified_payments)
            verified_interest = sum(payment.interest_amount for payment in verified_payments)
            
            # Calculate last paid date and amount (most recent verified payment)
            last_paid_date = None
            last_paid_amount = None
            if verified_payments:
                last_payment = max(verified_payments, key=lambda p: p.payment_date)
                last_paid_date = last_payment.payment_date.date()
                last_paid_amount = last_payment.amount
            
            # Calculate interest pending based on payment frequency
            interest_pending = accumulated_interest_monthly if loan.payment_frequency == 'monthly' else accumulated_interest_daily
            
            loan_data.append({
                'loan': loan,
                'daily_interest': daily_interest,
                'monthly_interest': monthly_interest,
                'accumulated_interest_daily': accumulated_interest_daily,
                'accumulated_interest_monthly': accumulated_interest_monthly,
                'interest_pending': interest_pending,
                'interest_data': interest_data,
                'pending_principal': pending_principal,
                'pending_interest': pending_interest,
                'pending_total': pending_total,
                'verified_principal': verified_principal,
                'verified_interest': verified_interest,
                'last_paid_date': last_paid_date,
                'last_paid_amount': last_paid_amount
            })
        
        return render_template('customer/all_loans.html',
                             loan_data=loan_data,
                             instance_name=instance_name)


    @app.route('/<instance_name>/customer/loan/<int:loan_id>')
    @login_required
    def customer_loan_detail(instance_name, loan_id):
        """Customer loan detail page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        # Check if loan belongs to current user and is active
        if loan.customer_id != current_user.id or not loan.is_active:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        # Calculate interest information
        daily_interest = calculate_daily_interest(loan.remaining_principal, loan.interest_rate)
        monthly_interest = calculate_monthly_interest(loan.remaining_principal, loan.interest_rate)
        interest_data = calculate_accumulated_interest(loan)
        accumulated_interest_daily = interest_data['daily']
        accumulated_interest_monthly = interest_data['monthly']
        
        # Calculate interest pending based on payment frequency (same logic as customer_all_loans)
        interest_pending = accumulated_interest_monthly if loan.payment_frequency == 'monthly' else accumulated_interest_daily
        
        # Get payment history
        payments = get_payment_query().filter_by(loan_id=loan_id).order_by(Payment.payment_date.desc()).all()
        
        # Calculate total interest paid for this loan
        total_interest_paid = sum(payment.interest_amount for payment in payments if payment.status == 'verified')
        
        # Calculate verified principal paid for this loan
        verified_principal = sum(payment.principal_amount for payment in payments if payment.status == 'verified')
        
        # Calculate pending amounts for this loan
        pending_principal = sum(payment.principal_amount for payment in payments if payment.status == 'pending')
        pending_interest = sum(payment.interest_amount for payment in payments if payment.status == 'pending')
        pending_total = pending_principal + pending_interest
        
        # Calculate days active
        days_active = (date.today() - loan.created_at.date()).days
        
        # Calculate previous principal and interest for each payment
        # We need to calculate what the principal and interest were before each payment
        payments_with_previous = []
        current_principal = loan.principal_amount
        current_interest_paid = 0
        
        # Process payments in chronological order (oldest first)
        payments_chronological = sorted(payments, key=lambda p: p.payment_date)
        
        for payment in payments_chronological:
            # Store the previous amounts before this payment
            previous_principal = current_principal
            previous_interest_paid = current_interest_paid
            
            # Update current amounts after this payment (only if verified)
            if payment.status == 'verified':
                current_principal -= payment.principal_amount
                current_interest_paid += payment.interest_amount
            
            # Add payment with previous amounts
            payments_with_previous.append({
                'payment': payment,
                'previous_principal': previous_principal,
                'previous_interest_paid': previous_interest_paid
            })
        
        # Reverse to show newest first
        payments_with_previous.reverse()
        
        # Calculate cashback for logged-in customer only
        from app_multi import CashbackTransaction, db_manager
        from decimal import Decimal
        
        session = db_manager.get_session_for_instance(instance_name)
        
        # Calculate cashback per payment for this customer
        payment_cashback_map = {}
        has_any_cashback = False
        for payment_data in payments_with_previous:
            payment = payment_data['payment']
            payment_cashback = session.query(
                db.func.sum(CashbackTransaction.points)
            ).filter_by(
                related_payment_id=payment.id,
                to_user_id=current_user.id
            ).scalar() or Decimal('0')
            payment_cashback_float = float(payment_cashback)
            payment_cashback_map[payment.id] = payment_cashback_float
            if payment_cashback_float > 0:
                has_any_cashback = True
        
        # Calculate total cashback for this loan for this customer
        loan_cashback_total = session.query(
            db.func.sum(CashbackTransaction.points)
        ).filter_by(
            related_loan_id=loan_id,
            to_user_id=current_user.id
        ).filter(
            CashbackTransaction.transaction_type.in_(['loan_interest_auto', 'loan_interest_manual'])
        ).scalar() or Decimal('0')
        
        return render_template('customer/loan_detail.html', 
                             loan=loan,
                             daily_interest=daily_interest,
                             monthly_interest=monthly_interest,
                             accumulated_interest_daily=accumulated_interest_daily,
                             accumulated_interest_monthly=accumulated_interest_monthly,
                             interest_pending=interest_pending,
                             interest_data=interest_data,
                             total_interest_paid=total_interest_paid,
                             verified_principal=verified_principal,
                             pending_principal=pending_principal,
                             pending_interest=pending_interest,
                             pending_total=pending_total,
                             payments=payments_with_previous,
                             days_active=days_active,
                             current_date=date.today(),
                             loan_cashback_total=float(loan_cashback_total),
                             payment_cashback_map=payment_cashback_map,
                             has_any_cashback=has_any_cashback,
                             instance_name=instance_name)


    @app.route('/<instance_name>/customer/loan/<int:loan_id>/edit-notes', methods=['GET', 'POST'])
    @login_required
    def customer_edit_notes(instance_name, loan_id):
        """Customer edit notes page for specific instance"""
        if instance_name not in VALID_INSTANCES:
            return redirect('/')
        
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard', instance_name=instance_name))
        
        loan = get_loan_query().filter_by(id=loan_id).first() or abort(404)
        
        # Check if loan belongs to current user
        if loan.customer_id != current_user.id:
            flash('Access denied')
            return redirect(url_for('customer_dashboard', instance_name=instance_name))
        
        if request.method == 'POST':
            customer_notes = request.form.get('customer_notes', '')
            loan.customer_notes = customer_notes
            commit_current_instance()
            
            flash('Notes updated successfully')
            return redirect(url_for('customer_loan_detail', instance_name=instance_name, loan_id=loan_id))
        
        return render_template('customer/edit_notes.html', 
                             loan=loan,
                             instance_name=instance_name)

